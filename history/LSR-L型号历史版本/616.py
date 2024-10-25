import json
import struct
import sys
import cv2
import time
import sqlite3
import os
import pygame
import serial.tools.list_ports
import re

from datetime import datetime
from zipfile import ZipFile
# import qdarkstyle
#
# from qt_material import apply_stylesheet
from paddleocr import PaddleOCR, draw_ocr

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QTabWidget, QWidget, QVBoxLayout, QPushButton, QLabel, QHBoxLayout, QTableWidget,
    QTableWidgetItem, QComboBox, QDialog, QMessageBox, QSpacerItem, QSizePolicy, QTextEdit, QLineEdit, QSlider,
    QCheckBox, QStyledItemDelegate, QGraphicsView, QGraphicsScene, QGraphicsPixmapItem
)
from PySide6.QtCore import Qt, QTimer, QThread, Signal, QCoreApplication, QDateTime, QObject
from PySide6.QtGui import QPixmap, QImage, QColor, QFont, QPalette
import numpy as np
from MechEye import Device
from dash.html import Figure
from matplotlib.font_manager import FontProperties
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

from mapping2depth import *
from mecheye_python_samples.source import Common
from now_demo_loop import CaptureAllData
import open3d as o3d
import torch
from pathlib import Path
from tqdm import tqdm
from PIL import Image
from openpyxl import Workbook, load_workbook


class DraculaColors:
    background = "#282a36"
    current_line = "#44475a"
    selection = "#44475a"
    foreground = "#f8f8f2"
    comment = "#6272a4"
    cyan = "#8be9fd"
    green = "#50fa7b"
    orange = "#ffb86c"
    pink = "#ff79c6"
    purple = "#bd93f9"
    red = "#ff5555"
    yellow = "#f1fa8c"


# 自定义委托类用于绘制下拉列表项
class ComboBoxDelegate(QStyledItemDelegate):
    def paint(self, painter, option, index):
        if index.data(Qt.UserRole):
            pixmap = QPixmap(index.data(Qt.UserRole))
            painter.drawPixmap(option.rect.x(), option.rect.y(), pixmap)
        else:
            super().paint(painter, option, index)


class CustomComboBox(QComboBox):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setItemDelegate(ComboBoxDelegate())


class YourClass(QObject):
    relay1_state_changed_signal = Signal(bool)
    relay2_state_changed_signal = Signal(bool)
    relay3_state_changed_signal = Signal(bool)
    relay4_state_changed_signal = Signal(bool)

    def __init__(self):
        super().__init__()
        # 初始化串口和其他参数
        available_ports = [port.device for port in serial.tools.list_ports.comports()]
        if available_ports:
            self.ser = serial.Serial("COM4", 9600, timeout=0.5)
            print(f"已连接至串口 {self.ser.port}, 波特率 9600")
            self.connected = True
        else:
            print("没有可用的串口")
            self.connected = False

    def send_command(self, relay_number, action_code):
        if not self.connected:
            print("串口未连接")
            return

        # 计算继电器的起始地址
        coil_address = relay_number - 1
        # 构造指令
        command = struct.pack(">BBHH", 1, 0x05, coil_address, action_code)
        crc = self.calculate_crc(command)
        command += struct.pack("<H", crc)

        try:
            self.ser.write(command)
            print(f"已发送指令: {command.hex()}")
        except Exception as e:
            print(f"发送指令时发生错误: {e}")

    @staticmethod
    def calculate_crc(data):
        crc = 0xFFFF
        for byte in data:
            crc ^= byte
            for _ in range(8):
                if crc & 1:
                    crc >>= 1
                    crc ^= 0xA001
                else:
                    crc >>= 1
        return crc

    def relay1_state_changed(self, state):
        print("Relay 1 State Changed:", state)
        self.send_command(1, 0xFF00 if state else 0x0000)

    def relay2_state_changed(self, state):
        print("Relay 2 State Changed:", state)
        self.send_command(2, 0xFF00 if state else 0x0000)

    def relay3_state_changed(self, state):
        print("Relay 3 State Changed:", state)
        self.send_command(3, 0xFF00 if state else 0x0000)

    def relay4_state_changed(self, state):
        print("Relay 4 State Changed:", state)
        self.send_command(4, 0xFF00 if state else 0x0000)


class ImageCaptureThread(QThread):
    stopCapture = Signal()
    resumeCapture = Signal()
    imagesCaptured = Signal(QImage, QImage)
    imagesCapturedAndSaved = Signal(str, str)
    analysisCompleted = Signal()
    detectionCompleted = Signal(object)
    depthvalue = Signal(str)
    depthvalue1 = Signal(str)
    imageCaptured = Signal(QImage, QImage)
    # 新增信号定义，用于传递LC最大值
    lcMaxUpdated = Signal(str)
    alarmStatusChanged = Signal(bool)

    def __init__(self, device, parent_window):

        super().__init__()
        self.device = device
        self.parent_window = parent_window
        self.is_running = True  # 标志来控制线程的运行状态
        self.model_cache = {}  # 字典用于存储加载的模型
        self.selected_model = None
        self.selected_model = "错缝"
        pygame.mixer.init()
        self.processing_functions = {
            '铝板': self.process_common_model,
            '栅格': self.process_grille,
            '冲孔板': self.process_common_model,
            '拉网': self.process_common_model,
            '铝条板': self.process_common_model,
            '栅格加网': self.process_common_model,
            '铝方通': self.process_aluminum_square_tube,
            '矿棉板': self.process_common_model,
            '勾搭铝板': self.process_common_model,
            '垂片': self.process_aluminum_square_tube
        }
        self.detected_labels = []
        # self.capture_interval = 10000
        self.imageCaptured.connect(self.on_images_captured)

    def run(self):
        timer = QTimer()
        timer.timeout.connect(self.capture_images)
        timer.start(self.capture_interval)  # 8000 毫秒（8秒）触发一次
        self.stopCapture.connect(lambda: timer.stop())
        self.resumeCapture.connect(lambda: timer.start(self.capture_interval))

        while self.is_running:
            QCoreApplication.processEvents()  # 处理事件，确保界面响应
            time.sleep(0.1)  # 避免CPU占用过高
        timer.stop()

    def set_capture_interval(self, interval):
        self.capture_interval = interval

    def set_selected_model(self, model_name):
        self.selected_model = model_name
        if model_name not in self.model_cache:
            self.model_cache[model_name] = self.load_model(model_name)

    def load_model(self, model_name):
        # 小电脑模型路径
        model_path = 'D:\PythonCode\yolov5'
        weight_paths = {
            '铝板': 'D:\\PythonCode\\yolov5\\runs\\train\\all\\511fanglvban.pt',
            '冲孔板': 'D:/PythonCode/yolov5/runs/train/all/523fanguang.pt',
            '拉网': 'D:/PythonCode/yolov5/runs/train/all/506heibai.pt',
            '铝条板': 'D:/PythonCode/yolov5/runs/train/all/522ltb.pt',
            '栅格加网': 'D:/PythonCode/yolov5/runs/train/all/shangejiawang.pt',
            '矿棉板': 'D:/PythonCode/yolov5/runs/train/all/kuangmian.pt',
            '栅格': 'D:/PythonCode/yolov5/runs/train/all/511lvfangtong.pt',
            '铝方通': 'D:/PythonCode/yolov5/runs/train/all/511lvfangtong.pt',
            '勾搭铝板': 'D:/PythonCode/yolov5/runs/train/all/cuowei.pt',
            '垂片': 'D:/PythonCode/yolov5/runs/train/all/511lvfangtong.pt'
        }
        weight_path = weight_paths.get(model_name, 'D:\PythonCode\yolov5/runs/train/all')
        return torch.hub.load(model_path, 'custom', weight_path, source='local')

    def capture_images(self):
        color_image = self.device.capture_color().data()
        depth_image = self.device.capture_depth().data()
        self.depth_image = depth_image

        self.height, self.width = color_image.shape[:2]
        if self.height != 1500 or self.width != 2000:
            print("change size!")
            color_image = cv2.resize(color_image, (2000, 1500), interpolation=cv2.INTER_AREA)
            depth_image = cv2.copyMakeBorder(depth_image, 70, 200, 90, 200, cv2.BORDER_CONSTANT, value=[255, 255, 255])
            depth_image = cv2.resize(depth_image, (2048, 1536), interpolation=cv2.INTER_AREA)
            self.depth_image = depth_image
        color_image_width, color_image_height = color_image.shape[1], color_image.shape[0]
        depth_image_width, depth_image_height = depth_image.shape[1], depth_image.shape[0]

        print(f"Color Image Resolution: {color_image_width} x {color_image_height}")
        print(f"Depth Image Resolution: {depth_image_width} x {depth_image_height}")

        depth_colormap = cv2.applyColorMap(cv2.convertScaleAbs(depth_image, alpha=0.03), cv2.COLORMAP_JET)
        self.depth_img = QImage(depth_colormap.data, depth_colormap.shape[1], depth_colormap.shape[0],
                                QImage.Format_RGB888)
        self.color_img = QImage(color_image.data, color_image.shape[1], color_image.shape[0], QImage.Format_RGB888)
        self.current_color_image = self.color_img
        self.current_depth_image = self.depth_img
        self.imagesCaptured.emit(self.color_img, self.depth_img)
        self.imagesCapturedAndSaved.emit(self.color_img, self.depth_img)

        if self.selected_model in self.processing_functions:
            detection_results = self.processing_functions[self.selected_model](color_image)
            self.detectionCompleted.emit(detection_results)
            self.imageCaptured.emit(self.color_img, self.depth_img)

    def calculate_vertical_distance(self, depth_value1, depth_value2):
        # 深度值以毫米为单位，直接相减得到垂直距离
        vertical_distance = abs(depth_value2 - depth_value1)
        return vertical_distance

    def process_common_model(self, color_frame):
        global depth_value_at_x2_y2, depth_value_at_x1_y1
        self.line_threshold = 67  # 基于您的图像尺寸和视野，您可能需要调整这个值
        self.column_threshold = 60
        # QTimer.singleShot(500, lambda: self.detectionCompleted.emit(detection_img))
        self.distance_to_object_data = {
            1200: (1200, 1200),  # 相机距离拍摄物1.2米，视野宽度1.2米，视野高度1米
            1300: (1300, 1070),  # 相机距离拍摄物1.3米，视野宽度1.3米，视野高度1.07米
            1400: (1400, 1130),  # 相机距离拍摄物1.4米，视野宽度1.4米，视野高度1.13米
            1500: (1500, 1200),  # 相机距离拍摄物1.5米，视野宽度1.5米，视野高度1.2米
            1600: (1600, 1280),  # 相机距离拍摄物1.6米，视野宽度1.6米，视野高度1.28米
            1700: (1700, 1360),  # 相机距离拍摄物1.7米，视野宽度1.7米，视野高度1.36米
            1800: (1800, 1440),  # 相机距离拍摄物1.8米，视野宽度1.8米，视野高度1.44米
            1900: (1900, 1520),  # 相机距离拍摄物1.9米，视野宽度1.9米，视野高度1.52米
            2000: (2000, 1600),  # 相机距离拍摄物2.0米，视野宽度2.0米，视野高度1.60米
            2100: (2100, 1680),  # 相机距离拍摄物2.1米，视野宽度2.1米，视野高度1.68米
            2200: (2200, 1760),  # 相机距离拍摄物2.2米，视野宽度2.2米，视野高度1.76米
            2300: (2300, 1840),  # 相机距离拍摄物2.3米，视野宽度2.3米，视野高度1.84米
            2400: (2400, 1920),  # 相机距离拍摄物2.4米，视野宽度2.4米，视野高度1.92米
            2500: (2500, 2000),  # 相机距离拍摄物2.5米，视野宽度2.5米，视野高度2.0米
            2600: (2600, 2080),  # 相机距离拍摄物2.6米，视野宽度2.6米，视野高度2.08米
            2700: (2700, 2160),  # 相机距离拍摄物2.7米，视野宽度2.7米，视野高度2.16米
            2800: (2800, 2240),  # 相机距离拍摄物2.8米，视野宽度2.8米，视野高度2.24米
            2900: (2900, 2320),  # 相机距离拍摄物2.9米，视野宽度2.9米，视野高度2.32米
            3000: (3000, 2400)  # 相机距离拍摄物3.0米，视野宽度3.0米，视野高度2.40米
        }
        if self.selected_model not in self.model_cache:
            # print(f"模型 {self.selected_model} 尚未加载")
            return

        model = self.model_cache[self.selected_model]
        results = model(color_frame)  # 假设model是已加载并准备好的模型实例
        ocr = PaddleOCR(use_angle_cls=True, lang="en")
        img = color_frame
        result = ocr.ocr(img, cls=True)

        if result is None:
            print("未检测到文本或OCR返回None")
            return

        # 迭代OCR结果
        for res in result:
            if res is None:
                continue  # 如果某个特定结果是None，则跳过
            for line in res:
                print("文字识别")
                print(line, "\tresult:", line[1][0])

        # 获取彩色图像宽高、深度图宽高
        image_width, image_height = color_frame.shape[1], color_frame.shape[0]  # 彩色图像的宽度和高度
        # 深度图：depth_frame , RGB图：color_frame
        # depth_frame = self.device.capture_depth().data()
        depth_frame = self.depth_image
        depth_height, depth_width = 1405 - 50, 1877 - 80  # 深度图像的宽度和高度
        depth_frame = depth_frame[50:1405, 80:1877]
        # 获取点云数据
        point_xyz_map = self.device.capture_point_xyz()
        point_xyz_data = point_xyz_map.data()

        fov_width_pixels = 0  # 视野范围，不用动
        fov_height_pixels = 0

        max_vertical_distance = {}
        horizontal_difference = 0
        # 在帧上绘制检测结果
        # 获取要分析的图像内容
        self.label_stats = {}
        self.detected_labels = []
        for *xyxy, conf, cls in results.xyxy[0]:
            mids = []
            vertical_distances = []
            x1, y1, x2, y2 = map(int, xyxy)
            if y2 - y1 <= 1 or x2 - x1 <= 1:
                # color_frame[y1:y2, x1:x2] = [255, 255, 255]
                # if y2 - y1 <= 1 or x2 - x1 <= 1:

                # print("检测到水平物体，跳过当前数据组")
                continue
            box_w, box_h = x2 - x1, y2 - y1
            if y1 < 50:
                y1 = 50
            if y2 < 50:
                y2 = 50
            # 更新标签统计字典
            label_name = results.names[int(cls)]
            if label_name not in self.label_stats:
                self.label_stats[label_name] = 1
            else:
                self.label_stats[label_name] += 1
            # 计算另外两个角的坐标
            x3, y3 = x1, y2  # 左下角
            x4, y4 = x2, y1  # 右上角
            # 调用映射坐标方法
            try:
                counts = mapping2depth(x1, y1, x2, y2)
            except:
                print("坐标框转换计算位置出错，跳过当前病害位置")
                continue

            for count in counts:
                try:
                    depth_x1, depth_y1, depth_x2, depth_y2, mean, shape = count
                    # cv2.rectangle(depth_frame, (depth_x1, depth_y1),
                    #               (depth_x2, depth_y2),
                    #               (0, 0, 255), 2)
                    # cv2.imshow("depth_frame", depth_frame)

                    # 调用映射坐标方法

                    # print(
                    #     f"Rectangular Coordinates: ({x1}, {y1}), ({x2}, {y2}), ({x3}, {y3}), ({x4}, {y4})")  # x1左上，x2右下，x3左下，x4右上
                    # 打印彩色图框位置信息
                    # print(f"Color Image Coordinates: ({x1}, {y1}) to ({x2}, {y2})")
                    # 打印深度图框位置信息
                    # print(f"Depth Image Coordinates: ({depth_x1}, {depth_y1}) to ({depth_x2}, {depth_y2})")
                    # 判断框的位置信息是不是在图像内
                    if 0 <= depth_x1 < depth_width and 0 <= depth_y1 < depth_height and \
                            0 <= depth_x2 < depth_width and 0 <= depth_y2 < depth_height:

                        # if self.height != 1500 or self.width != 2000:
                        #     x, y, depth_value_at_x1_y1 = point_xyz_data[int(depth_y1 + 59)][int(depth_x1 + 96)]
                        #     x, y, depth_value_at_x2_y2 = point_xyz_data[int(depth_y2 + 59)][int(depth_x2 + 96)]
                        # else:
                        if shape == 1 :
                            depth_y = (depth_y1 + depth_y2) / 2
                            x, y, depth_value_at_x1_y1 = point_xyz_data[int(depth_y + 50)][int(depth_x1 + 80)]
                            x, y, depth_value_at_x2_y2 = point_xyz_data[int(depth_y + 50)][int(depth_x2 + 80)]
                        elif shape == 0 :
                            depth_x = (depth_x1 + depth_x2) / 2
                            x, y, depth_value_at_x1_y1 = point_xyz_data[int(depth_y1 + 50)][int(depth_x + 80)]
                            x, y, depth_value_at_x2_y2 = point_xyz_data[int(depth_y2 + 50)][int(depth_x + 80)]
                        if depth_value_at_x1_y1 == 0 or depth_value_at_x2_y2 == 0:
                            # print("至少一个点的深度值为0，跳过当前数据组")
                            continue
                        distance_to_object = round(depth_value_at_x1_y1)

                        vertical_distance = self.calculate_vertical_distance(depth_value_at_x1_y1,
                                                                             depth_value_at_x2_y2)
                        vertical_distances.append(vertical_distance)
                        print(depth_value_at_x1_y1)
                        print(depth_value_at_x2_y2)
                        image_width = 2000
                        image_height = 1500
                        rounded_distance = round(distance_to_object, -2)  # 四舍五入到最接近的百位数
                        if rounded_distance in self.distance_to_object_data:
                            fov_width_pixels, fov_height_pixels = self.distance_to_object_data[rounded_distance]
                            # 其他处理...
                        else:
                            print(f"无法找到 {rounded_distance} 对应的键")

                        row_index = int(y1 / image_height * 10)
                        col_index = int(x1 / image_width * 10)
                        if (row_index, col_index) not in max_vertical_distance:
                            max_vertical_distance[(row_index, col_index)] = (
                                x1, y1, x2, y2, x3, y3, x4, y4, vertical_distance)
                        else:
                            existing_vertical_distance = max_vertical_distance[(row_index, col_index)][-1]
                            if vertical_distance > existing_vertical_distance:
                                max_vertical_distance[(row_index, col_index)] = (
                                    x1, y1, x2, y2, x3, y3, x4, y4, vertical_distance)
                        # if vertical_distance > 3:
                        cv2.rectangle(color_frame, (x1, y1), (x2, y2), (0, 0, 255), 2)


                    else:  # 框的位置信息不在图像内
                        # print("深度图像坐标超出范围，跳过当前数据组")
                        continue
                    try:

                        depth_frame = cv2.cvtColor(depth_frame, cv2.COLOR_BGR2GRAY)
                    except Exception as e:
                        print(f"转换深度图失败, 错误信息: {e}")
                    finally:
                        check_img = depth_frame[depth_y1:depth_y2, depth_x1:depth_x2]

                        w, h = check_img.shape[1], check_img.shape[0]
                        # 不用计算灰度阈值了，深度图的灰度阈值很清晰
                        # min = int(np.min(
                        #     check_img[int((w // 2) - 5):int((w // 2) + 5), int((h // 2) - 5):int((h // 2) + 5)]))
                        # 二值化，大于阈值得到1
                        ret, thresh = cv2.threshold(check_img, 100, 1, cv2.THRESH_BINARY)
                        # 矩阵之和，白一些的像素有多少
                        summ = np.sum(thresh)
                        # # 总像素减白像素除以宽/高 ，得到一行/一列黑像素有多少

                        # vertical_difference = ((w * h) - summ) // mean
                        # # print(vertical_difference)
                        if shape == 1 and summ != None:
                            row_sums = np.sum(thresh, axis=1)
                            # 总像素减白像素除以宽/高 ，得到一行/一列黑像素有多少
                            midd = ((w * h) - summ) // h
                            mids.append(midd)
                            # print("行和:", row_sums)
                        elif shape == 0 and summ != None:
                            # 计算每列的总和
                            row_sums = np.sum(thresh, axis=0)
                            # 总像素减白像素除以宽/高 ，得到一行/一列黑像素有多少
                            midd = ((w * h) - summ) // w
                            mids.append(midd)
                            # print("列和:", row_sums)
                        # vertical_list = [mean - i for i in row_sums]
                        # maxx, minn = max(vertical_list), min(vertical_list)
                except:
                    pass
            if len(mids) == 0:
                continue
            midd = max(mids)
            vertical_distance = max(vertical_distances)
            # vertical_difference = abs(y2 - y1)  # 左上角对应左下角的差值
            # vertical_distance1 = fov_height_pixels * maxx / image_height
            # vertical_distance2 = fov_height_pixels * minn / image_height
            vertical_distance3 = fov_height_pixels * midd / image_height
            detected_label = {
                "LC": vertical_distance,
                "LF": vertical_distance3,
            }
            self.detected_labels.append(detected_label)

            # 更新检测标签列表后，计算最大的LC值
            if self.detected_labels:
                max_lc = max(label['LC'] for label in self.detected_labels if 'LC' in label)
                self.lcMaxUpdated.emit(f"最大LC: {max_lc:.2f}")  # 发射信号，包含最大LC值
            else:
                self.lcMaxUpdated.emit("未检测到任何对象")  # 如果没有标签，发射此消息
            label = f'LC{vertical_distance:.2f}\nLF:{vertical_distance3:.2f}\n'
            if vertical_distance3 >= 100 or vertical_distance >= 3.0:
                self.alarmStatusChanged.emit(True)
                if vertical_distance3 >= 10:
                    print("触发报警！水平距离大于等于十公分。")
                elif vertical_distance >= 3.0:
                    print("触发报警！垂直距离大于等于三毫米。")

                    pygame.mixer.music.load("D:/PythonCode/sql/sp.mp3")  # 替换为你的音频文件路径
                    pygame.mixer.music.play()

                # if vertical_distance3 >= 10 or vertical_distance >= 3.0:
                #     if vertical_distance3 >= 10:
                #         print("触发报警！水平距离大于等于十公分。")
                #     elif vertical_distance >= 3.0:
                #         print("触发报警！垂直距离大于等于三毫米。")
                #
                #         pygame.mixer.music.load("D:/PycharmProjects/pythonProject/mp3/sp.mp3")  # 替换为你的音频文件路径
                #         pygame.mixer.music.play()
                print(f"Rectangular Coordinates: ({x1}, {y1}), ({x2}, {y2}), ({x3}, {y3}), ({x4}, {y4})")


            # cv2.putText(color_frame, label, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 0, 255), 2)
            else:
                self.alarmStatusChanged.emit(False)
            lines = label.split('\n')
            y = y1 + 30
            for label in lines:
                cv2.putText(color_frame, label, (x1 + (box_w // 2), y), cv2.FONT_HERSHEY_SIMPLEX, 1.0,
                            (0, 255, 0), 2)
                y += 30

            # cv2.putText(color_frame, label, (x1 + (box_w // 2), y1 + (box_h // 2)), cv2.FONT_HERSHEY_SIMPLEX, 1.0,
            #             (0, 0, 255), 2)
        detection_img = QImage(color_frame.data, image_width, image_height, QImage.Format_RGB888)
        return detection_img

    def process_grille(self, color_frame):
        self.line_threshold = 67  # 基于您的图像尺寸和视野，您可能需要调整这个值
        self.column_threshold = 60
        # QTimer.singleShot(500, lambda: self.detectionCompleted.emit(detection_img))
        self.distance_to_object_data = {
            1200: (1200, 1200),  # 相机距离拍摄物1.2米，视野宽度1.2米，视野高度1米
            1300: (1300, 1070),  # 相机距离拍摄物1.3米，视野宽度1.3米，视野高度1.07米
            1400: (1400, 1130),  # 相机距离拍摄物1.4米，视野宽度1.4米，视野高度1.13米
            1500: (1500, 1200),  # 相机距离拍摄物1.5米，视野宽度1.5米，视野高度1.2米
            1600: (1600, 1280),  # 相机距离拍摄物1.6米，视野宽度1.6米，视野高度1.28米
            1700: (1700, 1360),  # 相机距离拍摄物1.7米，视野宽度1.7米，视野高度1.36米
            1800: (1800, 1440),  # 相机距离拍摄物1.8米，视野宽度1.8米，视野高度1.44米
            1900: (1900, 1520),  # 相机距离拍摄物1.9米，视野宽度1.9米，视野高度1.52米
            2000: (2000, 1600),  # 相机距离拍摄物2.0米，视野宽度2.0米，视野高度1.60米
            2100: (2100, 1680),  # 相机距离拍摄物2.1米，视野宽度2.1米，视野高度1.68米
            2200: (2200, 1760),  # 相机距离拍摄物2.2米，视野宽度2.2米，视野高度1.76米
            2300: (2300, 1840),  # 相机距离拍摄物2.3米，视野宽度2.3米，视野高度1.84米
            2400: (2400, 1920),  # 相机距离拍摄物2.4米，视野宽度2.4米，视野高度1.92米
            2500: (2500, 2000),  # 相机距离拍摄物2.5米，视野宽度2.5米，视野高度2.0米
            2600: (2600, 2080),  # 相机距离拍摄物2.6米，视野宽度2.6米，视野高度2.08米
            2700: (2700, 2160),  # 相机距离拍摄物2.7米，视野宽度2.7米，视野高度2.16米
            2800: (2800, 2240),  # 相机距离拍摄物2.8米，视野宽度2.8米，视野高度2.24米
            2900: (2900, 2320),  # 相机距离拍摄物2.9米，视野宽度2.9米，视野高度2.32米
            3000: (3000, 2400)  # 相机距离拍摄物3.0米，视野宽度3.0米，视野高度2.40米
        }
        if self.selected_model not in self.model_cache:
            print(f"模型 {self.selected_model} 尚未加载")
            return
        model = self.model_cache[self.selected_model]
        model.conf = 0.2  # 根据需要调整置信度阈值
        model.iou = 0.0  # 根据需要调整NMS IOU阈值
        image_width = 2000
        image_height = 1500
        # 进行检测
        results = model(color_frame)
        # 获取彩色图像宽高、深度图宽高
        image_width, image_height = color_frame.shape[1], color_frame.shape[0]  # 彩色图像的宽度和高度
        # 深度图：depth_frame , RGB图：color_frame
        # depth_frame = self.device.capture_depth().data()
        depth_frame = self.depth_image
        depth_height, depth_width = 1405 - 50, 1877 - 80  # 深度图像的宽度和高度
        depth_frame = depth_frame[50:1405, 80:1877]
        # 获取点云数据
        point_xyz_map = self.device.capture_point_xyz()
        point_xyz_data = point_xyz_map.data()
        fov_width_pixels = 0  # 视野范围，不用动
        fov_height_pixels = 0

        max_vertical_distance = {}
        horizontal_difference = 0
        # 在帧上绘制检测结果
        # 获取要分析的图像内容
        self.label_stats = {}
        self.detected_labels = []
        for *xyxy, conf, cls in results.xyxy[0]:
            if cls == 1 or cls == "cuofengg":
                x1, y1, x2, y2 = map(int, xyxy)
                if y2 - y1 <= 1 or x2 - x1 <= 1 or x2 < 200 or y2 > 1400 or y1 < 85:
                    print("检测到水平物体，跳过当前数据组")
                    continue
                x1, x2 = min(x1, x2), max(x1, x2)
                y1, y2 = min(y1, y2), max(y1, y2)
                box_w, box_h = x2 - x1, y2 - y1
                if y1 < 50:
                    y1 = 50
                if y2 < 50:
                    y2 = 50
                # 更新标签统计字典
                label_name = results.names[int(cls)]
                if label_name not in self.label_stats:
                    self.label_stats[label_name] = 1
                else:
                    self.label_stats[label_name] += 1
                # 计算另外两个角的坐标
                x3, y3 = x1, y2  # 左下角
                x4, y4 = x2, y1  # 右上角
                # 映射
                b_depth_x1 = int(x1 * depth_width / image_width)
                b_depth_y1 = int(y1 * depth_height / image_height)
                b_depth_x2 = int(x2 * depth_width / image_width)
                b_depth_y2 = int(y2 * depth_height / image_height)

                is_horizontal = abs(x2 - x1) > abs(y2 - y1)
                if is_horizontal:
                    c_depth_x1, c_depth_y1, c_depth_x2, c_depth_y2 = b_depth_x1, b_depth_y1 + (
                            box_h // 4), b_depth_x2, b_depth_y2 - (box_h // 4)
                    mean = c_depth_y2 - c_depth_y1
                    shape = 0
                    # 两个点计算直线度
                    z1y = (c_depth_y1 + c_depth_y2) // 2
                    zh = int((y2 - y1) * 0.7)
                    z1 = c_depth_x1, z1y
                    z2 = c_depth_x2, z1y
                    # # 计算获取深度z值的点竖框x轴的中心点
                    # d_depth_x1, d_depth_y1, d_depth_x2, d_depth_y2 = c_depth_x1, (c_depth_y1+c_depth_y2) // 2, c_depth_x2, (c_depth_y1+c_depth_y2) // 2
                else:
                    c_depth_x1, c_depth_y1, c_depth_x2, c_depth_y2 = b_depth_x1 + (
                            box_w // 4), b_depth_y1, b_depth_x2 - (box_w // 4), b_depth_y2
                    mean = c_depth_x2 - c_depth_x1
                    shape = 1
                    # 两个点计算直线度
                    z1x = (c_depth_x1 + c_depth_x2) // 2
                    zw = int((x2 - x1) * 0.7)
                    z1 = z1x, c_depth_y1, z1x + int((c_depth_x2 - c_depth_x1) * 0.7), c_depth_y1
                    z2 = z1x, c_depth_y2, z1x + int((c_depth_x2 - c_depth_x1) * 0.7), c_depth_y2
                    # # 计算获取深度z值的点竖框y轴的中心点
                    # d_depth_x1, d_depth_y1, d_depth_x2, d_depth_y2 = (c_depth_x1 + c_depth_x2) // 2, c_depth_y1, (c_depth_x1 + c_depth_x2) // 2, c_depth_y2
                # 调整框在深度图上的位置
                if c_depth_y2 < 1250:
                    if c_depth_y2 <= 857:
                        # print("c_x1+14,c_x2+14")
                        c_depth_x1, c_depth_x2 = c_depth_x1 + 14, c_depth_x2 + 14
                    elif 875 < c_depth_y2 and b_depth_x2 < 358:
                        # print("c_y1 - 17, c_y2 - 17")
                        c_depth_y1, c_depth_y2 = c_depth_y1 - 17, c_depth_y2 - 17
                        if c_depth_y2 > 1100:
                            # print("c_x1, c_x2 = c_x1 - 7, c_x2 - 7")
                            c_depth_x1, c_depth_x2 = c_depth_x1 - 7, c_depth_x2 - 7
                    elif c_depth_x2 > 1000 and c_depth_y2 > 931:
                        # print("c_y1 + 14, c_y2 + 14")
                        c_depth_y1, c_depth_y2 = c_depth_y1 + 14, c_depth_y2 + 14
                    if is_horizontal:
                        # 计算获取深度z值的点竖框x轴的中心点
                        d_depth_x1, d_depth_y1, d_depth_x2, d_depth_y2 = c_depth_x1, (
                                                                                             c_depth_y1 + c_depth_y2) // 2, c_depth_x2, (
                                                                                             c_depth_y1 + c_depth_y2) // 2
                    else:
                        # 计算获取深度z值的点竖框y轴的中心点
                        d_depth_x1, d_depth_y1, d_depth_x2, d_depth_y2 = (
                                                                                 c_depth_x1 + c_depth_x2) // 2, c_depth_y1, (
                                                                                 c_depth_x1 + c_depth_x2) // 2, c_depth_y2
                    print(
                        f"Rectangular Coordinates: ({x1}, {y1}), ({x2}, {y2}), ({x3}, {y3}), ({x4}, {y4})")  # x1左上，x2右下，x3左下，x4右上
                    # 打印彩色图框位置信息
                    print(f"Color Image Coordinates: ({x1}, {y1}) to ({x2}, {y2})")
                    # 打印深度图框位置信息
                    print(f"Depth Image Coordinates: ({c_depth_x1}, {c_depth_y1}) to ({c_depth_x2}, {c_depth_y2})")
                    # 判断框的位置信息是不是在图像内
                    if 0 <= c_depth_x1 < depth_width and 0 <= c_depth_y1 < depth_height and \
                            0 <= c_depth_x2 < depth_width and 0 <= c_depth_y2 < depth_height:
                        if self.height != 1500 or self.width != 2000:
                            x, y, depth_value_at_x1_y1 = point_xyz_data[int(d_depth_y1 + 59)][int(d_depth_x1 + 96)]
                            x, y, depth_value_at_x2_y2 = point_xyz_data[int(d_depth_y2 + 59)][int(d_depth_x2 + 96)]
                        else:
                            # 获取深度值，此处进行计算
                            x, y, depth_value_at_x1_y1 = point_xyz_data[d_depth_y1 + 50, d_depth_x1 + 80]
                            x, y, depth_value_at_x2_y2 = point_xyz_data[d_depth_y2 + 50, d_depth_x2 + 80]
                        distance_to_object = round(depth_value_at_x1_y1)

                        vertical_distance_int = self.calculate_vertical_distance(depth_value_at_x1_y1,
                                                                                 depth_value_at_x2_y2)
                        vertical_distance = f"LC{vertical_distance_int: .2f}"
                        if depth_value_at_x1_y1 < 1 or depth_value_at_x2_y2 < 1:
                            vertical_distance = "loss"
                        print(depth_value_at_x1_y1)
                        print(depth_value_at_x2_y2)
                        image_width = 2000
                        image_height = 1500
                        rounded_distance = round(distance_to_object, -2)  # 四舍五入到最接近的百位数
                        if rounded_distance in self.distance_to_object_data:
                            fov_width_pixels, fov_height_pixels = self.distance_to_object_data[rounded_distance]
                            # 其他处理...
                        else:
                            print(f"无法找到 {rounded_distance} 对应的键")

                        row_index = int(y1 / image_height * 10)
                        col_index = int(x1 / image_width * 10)
                        if (row_index, col_index) not in max_vertical_distance:
                            max_vertical_distance[(row_index, col_index)] = (
                                x1, y1, x2, y2, x3, y3, x4, y4, vertical_distance)
                        else:
                            existing_vertical_distance = max_vertical_distance[(row_index, col_index)][-1]
                            if vertical_distance > existing_vertical_distance:
                                max_vertical_distance[(row_index, col_index)] = (
                                    x1, y1, x2, y2, x3, y3, x4, y4, vertical_distance)

                        cv2.rectangle(color_frame, (x1, y1), (x2, y2), (0, 0, 255), 2)


                    else:  # 框的位置信息不在图像内
                        print("深度图像坐标超出范围，跳过当前数据组")
                        continue
                    try:
                        # 转换灰度图 直接调用深度图，深度图就是灰度图
                        # check_img = cv2.cvtColor(check_img, cv2.COLOR_BGR2GRAY)
                        # 计算区域内最小值(最黑)

                        depth_frame = cv2.cvtColor(depth_frame, cv2.COLOR_BGR2GRAY)
                    except Exception as e:
                        print(f"转换深度图失败, 错误信息: {e}")
                    finally:
                        check_img = depth_frame[c_depth_y1:c_depth_y2, c_depth_x1:c_depth_x2]

                        w, h = check_img.shape[1], check_img.shape[0]
                        # 不用计算灰度阈值了，深度图的灰度阈值很清晰
                        # min = int(np.min(
                        #     check_img[int((w // 2) - 5):int((w // 2) + 5), int((h // 2) - 5):int((h // 2) + 5)]))
                        # 二值化，大于阈值得到1
                        ret, thresh = cv2.threshold(check_img, 100, 1, cv2.THRESH_BINARY)
                        # 矩阵之和，白一些的像素有多少
                        summ = np.sum(thresh)
                        # # 总像素减白像素除以宽/高 ，得到一行/一列黑像素有多少

                        # vertical_difference = ((w * h) - summ) // mean
                        # # print(vertical_difference)
                        if shape == 1 and summ != None:
                            row_sums = np.sum(thresh, axis=1)
                            # 总像素减白像素除以宽/高 ，得到一行/一列黑像素有多少
                            midd = ((w * h) - summ) // w
                            z = 0
                            zz = [point_xyz_data[z1[1], z1[0]], point_xyz_data[z2[1], z2[0]]]
                            for i in range(zw):
                                z1x, z1y, z1z = point_xyz_data[int(z1[1]) + 50, int(z1[0] + i) + 80]
                                z2x, z2y, z2z = point_xyz_data[int(z2[1]) + 50, int(z2[0] + i) + 80]
                                # 假设zz, z1z, 和 z2z已经定义且是NumPy数组
                                condition1 = np.abs(zz[0] - z1z) > 500
                                condition2 = np.abs(zz[1] - z2z) > 500
                                c1 = zz[0] + 10
                                print("z1", z1z, "z2", z2z)

                                # 检查是否至少有一个元素满足条件
                                if condition1.any() or condition2.any():
                                    zc = abs(z1z - z2z)
                                    if zc > 150:
                                        if (zc < c1).any():
                                            # print(f"{zc}")
                                            z += 1
                            # print("行和:", row_sums)
                        elif shape == 0 and summ != None:
                            # 计算每列的总和
                            row_sums = np.sum(thresh, axis=0)
                            # 总像素减白像素除以宽/高 ，得到一行/一列黑像素有多少
                            midd = ((w * h) - summ) // h
                            z = 0
                            zz = [point_xyz_data[z1[1], z1[0]], point_xyz_data[z2[1], z2[0]]]
                            for i in range(zh):
                                z1x, z1y, z1z = point_xyz_data[int(z1[1] + i) + 50, int(z1[0]) + 80]
                                z2x, z2y, z2z = point_xyz_data[int(z2[1] + i) + 50, int(z2[0]) + 80]
                                # 假设zz, z1z, 和 z2z已经定义且是NumPy数组
                                condition1 = np.abs(zz[0] - z1z) > 500
                                condition2 = np.abs(zz[1] - z2z) > 500
                                c1 = zz[0] - 10
                                # print(f"z1\t,{z1x}\t, {z1y}\t, {z1z}\t, z2\t, {z2x}\t, {z2y}\t, {z2z}\t")
                                # print(f"z1,{z1z}\t,z2z,{z2z}")

                                # 检查是否至少有一个元素满足条件
                                if condition1.any() or condition2.any():
                                    zc = abs(z1z - z2z)
                                    if zc > 150:
                                        if (zc < c1).any():
                                            # print(f"{zc}")
                                            z += 1
                            # print("列和:", row_sums)
                        else:
                            print("框的位置出错")
                            midd = 0
                            z = 0
                        # vertical_list = [mean - i for i in row_sums]
                        # maxx, minn = max(vertical_list), min(vertical_list)
                        # cv2.rectangle(depth_frame, (int(c_depth_x1 - 3), int(c_depth_y1 - 3)),
                        #               (int(c_depth_x2 + 3), int(c_depth_y2 + 3)),
                        #               (0, 0, 255), 2)

                        # vertical_difference = abs(y2 - y1)  # 左上角对应左下角的差值
                    # vertical_distance1 = fov_height_pixels * maxx / image_height
                    # vertical_distance2 = fov_height_pixels * minn / image_height
                    vertical_distance3 = fov_height_pixels * midd / image_height
                    vertical_distance4 = fov_height_pixels * z / image_height
                    detected_label = {
                        "LC": vertical_distance,
                        "LF": vertical_distance3,
                        "CF": vertical_distance4,
                    }
                    self.detected_labels.append(detected_label)
                    if self.detected_labels:
                        max_lc = max(label['LC'] for label in self.detected_labels if 'LC' in label)
                        max_lc1 = f"最大LC: {max_lc}"
                        self.lcMaxUpdated.emit(max_lc1)  # 发射信号，包含最大LC值
                    else:
                        self.lcMaxUpdated.emit("未检测到任何对象")  # 如果没有标签，发射此消息
                    label = f'LC:{vertical_distance}\nLF:{vertical_distance3:.2f}\nCF:{vertical_distance4:.2f}\n'
                    if vertical_distance3 >= 10 or vertical_distance_int >= 3.0:
                        self.alarmStatusChanged.emit(True)
                        if vertical_distance3 >= 10:
                            print("触发报警！水平距离大于等于十公分。")
                        elif vertical_distance_int >= 3.0:
                            print("触发报警！垂直距离大于等于三毫米。")

                            pygame.mixer.music.load("D:/PythonCode/sql/sp.mp3")  # 替换为你的音频文件路径
                            pygame.mixer.music.play()

                        print(f"Rectangular Coordinates: ({x1}, {y1}), ({x2}, {y2}), ({x3}, {y3}), ({x4}, {y4})")
                        # 文本位置在框中间
                        # cv2.putText(color_frame, label, (x1 + (box_w // 2), y1 + (box_h // 2)), cv2.FONT_HERSHEY_SIMPLEX,
                        #             1.0, (0, 0, 255), 2)

                    else:
                        self.alarmStatusChanged.emit(False)
                    lines = label.split('\n')
                    y = y1 + 30
                    for label in lines:
                        cv2.putText(color_frame, label, (x2, y), cv2.FONT_HERSHEY_SIMPLEX, 1.0,
                                    (0, 255, 0), 2)
                        y += 30
                    # cv2.putText(color_frame, label, (x1 + (box_w // 2), y1 + (box_h // 2)),
                    #             cv2.FONT_HERSHEY_SIMPLEX, 1.0,
                    #             (0, 0, 255), 2)
                    # cv2.putText(color_frame, label, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 0, 255), 2)
        # 处理检测结果
        # self.process_detections(results, color_frame)
        # 更新UI显示
        # self.display_image(color_frame)
        detection_img = QImage(color_frame.data, image_width, image_height, QImage.Format_RGB888)
        return detection_img

    def process_aluminum_square_tube(self, color_frame):
        self.line_threshold = 67  # 基于您的图像尺寸和视野，您可能需要调整这个值
        self.column_threshold = 60
        # QTimer.singleShot(500, lambda: self.detectionCompleted.emit(detection_img))
        self.distance_to_object_data = {
            1200: (1200, 1200),  # 相机距离拍摄物1.2米，视野宽度1.2米，视野高度1米
            1300: (1300, 1070),  # 相机距离拍摄物1.3米，视野宽度1.3米，视野高度1.07米
            1400: (1400, 1130),  # 相机距离拍摄物1.4米，视野宽度1.4米，视野高度1.13米
            1500: (1500, 1200),  # 相机距离拍摄物1.5米，视野宽度1.5米，视野高度1.2米
            1600: (1600, 1280),  # 相机距离拍摄物1.6米，视野宽度1.6米，视野高度1.28米
            1700: (1700, 1360),  # 相机距离拍摄物1.7米，视野宽度1.7米，视野高度1.36米
            1800: (1800, 1440),  # 相机距离拍摄物1.8米，视野宽度1.8米，视野高度1.44米
            1900: (1900, 1520),  # 相机距离拍摄物1.9米，视野宽度1.9米，视野高度1.52米
            2000: (2000, 1600),  # 相机距离拍摄物2.0米，视野宽度2.0米，视野高度1.60米
            2100: (2100, 1680),  # 相机距离拍摄物2.1米，视野宽度2.1米，视野高度1.68米
            2200: (2200, 1760),  # 相机距离拍摄物2.2米，视野宽度2.2米，视野高度1.76米
            2300: (2300, 1840),  # 相机距离拍摄物2.3米，视野宽度2.3米，视野高度1.84米
            2400: (2400, 1920),  # 相机距离拍摄物2.4米，视野宽度2.4米，视野高度1.92米
            2500: (2500, 2000),  # 相机距离拍摄物2.5米，视野宽度2.5米，视野高度2.0米
            2600: (2600, 2080),  # 相机距离拍摄物2.6米，视野宽度2.6米，视野高度2.08米
            2700: (2700, 2160),  # 相机距离拍摄物2.7米，视野宽度2.7米，视野高度2.16米
            2800: (2800, 2240),  # 相机距离拍摄物2.8米，视野宽度2.8米，视野高度2.24米
            2900: (2900, 2320),  # 相机距离拍摄物2.9米，视野宽度2.9米，视野高度2.32米
            3000: (3000, 2400)  # 相机距离拍摄物3.0米，视野宽度3.0米，视野高度2.40米
        }
        if self.selected_model not in self.model_cache:
            print(f"模型 {self.selected_model} 尚未加载")
            return
        model = self.model_cache[self.selected_model]
        model.conf = 0.2  # 根据需要调整置信度阈值
        model.iou = 0.0  # 根据需要调整NMS IOU阈值
        image_width = 2000
        image_height = 1500
        # 进行检测
        results = model(color_frame)
        # 处理检测结果
        self.process_detections(results, color_frame)
        # 更新UI显示
        # self.display_image(color_frame)
        detection_img = QImage(color_frame.data, image_width, image_height, QImage.Format_RGB888)
        return detection_img

    def draw_boxes(self, image, boxes, distances, y_diffs=None, depth_diffs=None, color=(0, 255, 0), thickness=2,
                   font_scale=0.5, font_thickness=2):
        for i, box in enumerate(boxes):
            x1, y1, x2, y2 = box[:4]
            cv2.rectangle(image, (int(x1), int(y1)), (int(x2), int(y2)), color, thickness)
            info_texts = []
            if i < len(distances):
                info_texts.append(f"LF: {distances[i]:.2f}\n")
            if y_diffs and i < len(y_diffs):
                info_texts.append(f"CF: {y_diffs[i]:.2f}\n")
            if depth_diffs and i < len(depth_diffs):
                info_texts.append(f"LC: {depth_diffs[i]:.2f}\n")

            text = "".join(info_texts)
            lines = text.split('\n')
            y = int(y1) - 10
            for label in lines:
                cv2.putText(image, label, (int(x2) + 100, y), cv2.FONT_HERSHEY_SIMPLEX, font_scale, color,
                            font_thickness)
                y += 30
            # cv2.putText(image, text, (int(x1) + 10, int(y1) - 10), cv2.FONT_HERSHEY_SIMPLEX, font_scale, color,
            #             font_thickness)

    def draw_boxes_for_type(self, boxes, color_frame, color=(0, 255, 0), thickness=2):
        for box in boxes:
            x1, y1, x2, y2 = box[:4]
            cv2.rectangle(color_frame, (int(x1), int(y1)), (int(x2), int(y2)), color, thickness)

    def process_detections(self, results, color_frame):
        '''
        :param results: 框的信息
        :param color_frame: 彩色图片
        :return:
        '''
        image_width_px = 2000  # 图像宽度，单位：像素
        image_height_px = 1500  # 图像高度，单位：像素

        detections = results.xyxy[0].numpy()  # 转换结果为numpy数组

        cuofeng_class_id = 0  # 假设的类别ID，需要用实际的替换
        cuofengg_class_id = 1  # 假设的另一个类别ID

        cuofengg_boxes = detections[detections[:, 5] == cuofengg_class_id]  # 筛选cuofengg类型的边界框
        boxes_type_1 = detections[detections[:, 5] == cuofeng_class_id]  # 筛选cuofeng类型的边界框
        # depth_frame1 = self.device.capture_depth().data()
        depth_frame1 = self.depth_image
        # 读取深度图
        depth_frame = depth_frame1[50:1405, 80:1877]
        # 深度图的分辨率
        depth_height, depth_width = 1405 - 50, 1877 - 80
        # 获取点云数据
        point_xyz_map = self.device.capture_point_xyz()
        point_xyz_data = point_xyz_map.data()
        self.fov_width_pixels = 0
        self.fov_height_pixels = 0
        self.label_stats = {}

        for *xyxy, conf, cls in results.xyxy[0]:
            # x1, y1, x2, y2 = map(int, xyxy)
            # print("yuanshi", results.xyxy[0])

            # 更新标签统计字典
            label_name = results.names[int(cls)]
            if label_name not in self.label_stats:
                self.label_stats[label_name] = 1
            else:
                self.label_stats[label_name] += 1

        # 打印每个边界框的四个顶点坐标
        for detection in detections:
            x1, y1, x2, y2, conf, class_id = detection[:6]
            print(f"边界框类别ID {class_id}:")
            print(f"    左上角 (x1, y1): ({x1}, {y1})")
            print(f"    右上角 (x2, y1): ({x2}, {y1})")
            print(f"    左下角 (x1, y2): ({x1}, {y2})")
            print(f"    右下角 (x2, y2): ({x2}, {y2})")
            # depth_height, depth_width = 1405 - 50, 1877 - 80
            depth_x1 = int(x1 * depth_width / image_width_px)
            depth_y1 = int(y1 * depth_height / image_height_px)
            depth_x2 = int(x2 * depth_width / image_width_px)
            depth_y2 = int(y2 * depth_height / image_height_px)
            depth_value_at_x1_y1 = depth_frame[depth_y1, depth_x1]
            # depth_value_at_x2_y2 = depth_frame[depth_y2, depth_x2]
            distance_to_object = round(depth_value_at_x1_y1)
            rounded_distance = round(distance_to_object, -2)  # 四舍五入到最接近的百位数

            if rounded_distance in self.distance_to_object_data:
                self.fov_width_pixels, self.fov_height_pixels = self.distance_to_object_data[rounded_distance]
            else:
                print(f"无法找到 {rounded_distance} 对应的键")
            self.pixel_size_width_m = self.fov_width_pixels / image_width_px  # 每像素代表的宽度，单位：米
            self.pixel_size_height_m = self.fov_height_pixels / image_height_px  # 每像素代表的高度，单位：米

        def calculate_pixel_difference(point1, point2):
            return abs(point1[0] - point2[0]), abs(point1[1] - point2[1])

        # 定义一个函数来计算框的中心点
        def calculate_box_center(box):
            return ((box[0] + box[2]) / 2, (box[1] + box[3]) / 2), box[2] - box[0]

        for cuofengg_box in cuofengg_boxes:
            # depth_frame = self.device.capture_depth().data()

            # depth_height, depth_width = 1405 - 50, 1877 - 80
            inside_cuofeng_boxes = detections[(detections[:, 5] == cuofeng_class_id) &
                                              (detections[:, 0] >= cuofengg_box[0]) &
                                              (detections[:, 1] >= cuofengg_box[1]) &
                                              (detections[:, 2] <= cuofengg_box[2]) &
                                              (detections[:, 3] <= cuofengg_box[3])]
            inside_cuofeng_boxes = sorted(inside_cuofeng_boxes, key=lambda x: x[0])  # 按x坐标排序
            print(f"k:{inside_cuofeng_boxes}")
            for i in range(len(inside_cuofeng_boxes) - 1):
                # print("xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx\n",
                #       f"kuang1:{inside_cuofeng_boxes[i]}\tkuang2:{inside_cuofeng_boxes[i + 1]}")

                center1, box_h = calculate_box_center(inside_cuofeng_boxes[i])
                center2, box2_h = calculate_box_center(inside_cuofeng_boxes[i + 1])
                # print("xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx\n",
                #       f"kuang1:{center1, box_h}\tkuang2:{center2, box2_h}")

                x1, y1 = center1
                x2, y2 = center2
                x1, x2 = min(x1, x2), max(x1, x2)
                y1, y2 = min(y1, y2), max(y1, y2)
                is_horizontal = abs(x2 - x1) > abs(y2 - y1)

                if is_horizontal:
                    x1, y1, x2, y2 = x1, y1 - (box_h // 4), x2, y2 + (box2_h // 4)
                else:
                    x1, y1, x2, y2 = x1 - (box_h // 4), y1, x2 + (box2_h // 4), y2

                # # 深度图的分辨率
                # depth_height, depth_width = 1405 - 50, 1877 - 80
                # RGB图的分辨率
                # image_width, image_height = 2000, 1500
                # 映射到深度图
                c_depth_x1 = int(x1 * depth_width / image_width_px)
                c_depth_y1 = int(y1 * depth_height / image_height_px)
                c_depth_x2 = int(x2 * depth_width / image_width_px)
                c_depth_y2 = int(y2 * depth_height / image_height_px)

                if c_depth_y2 < 1250:
                    if c_depth_y2 <= 857:
                        # print("c_depth_x1+14,depth_x2+14")
                        c_depth_x1, c_depth_x2 = c_depth_x1 + 18, c_depth_x2 + 18
                    if c_depth_x1 > 1000 and 857 < c_depth_y1:
                        c_depth_x1, c_depth_x2 = c_depth_x1 + 18, c_depth_x2 + 18
                    if 875 < c_depth_y1 and c_depth_x1 < 358:
                        # print("depth_y1 - 17, c_depth_y2 - 17")
                        c_depth_y1, c_depth_y2 = c_depth_y1 - 17, c_depth_y2 - 17
                        if c_depth_y1 > 1100:
                            # print("c_depth_x1, c_depth_x2 = c_depth_x1 - 7, c_depth_x2 - 7")
                            c_depth_x1, c_depth_x2 = c_depth_x1 - 7, c_depth_x2 - 7
                    if c_depth_x1 > 1000 and c_depth_y1 > 931:
                        # print("depth_y1 + 14, c_depth_y2 + 14")
                        c_depth_y1, c_depth_y2 = c_depth_y1 + 14, c_depth_y2 + 14
                # cv2.rectangle(depth_frame, (c_depth_x1, c_depth_y1),
                #               (c_depth_x2, c_depth_y2),
                #               (0, 0, 255), 2)
                # 截取要计算的位置
                try:
                    # 转换灰度图 直接调用深度图，深度图就是灰度图
                    # check_img = cv2.cvtColor(check_img, cv2.COLOR_BGR2GRAY)
                    # 计算区域内最小值(最黑)

                    depth_frame = cv2.cvtColor(depth_frame, cv2.COLOR_BGR2GRAY)
                except Exception as e:
                    print(f"转换深度图失败, 错误信息: {e}")
                # depth_fram = cv2.cvtColor(depth_frame, cv2.COLOR_BGR2GRAY)
                check_img = depth_frame[c_depth_y1:c_depth_y2, c_depth_x1:c_depth_x2]
                w, h = check_img.shape[1], check_img.shape[0]
                # 二值化，大于阈值得到1
                ret, thresh = cv2.threshold(check_img, 100, 1, cv2.THRESH_BINARY)
                # 矩阵之和，白一些的像素有多少
                # print(thresh)
                summ = np.sum(thresh)
                # is_horizontal = abs(c_depth_x2 - c_depth_x1) > abs(c_depth_y2 - c_depth_y1)
                if is_horizontal:
                    # 横向框
                    right_top_point = [inside_cuofeng_boxes[i][2], inside_cuofeng_boxes[i][1]]  # 左边框的右上角顶点
                    left_top_point = [inside_cuofeng_boxes[i + 1][0], inside_cuofeng_boxes[i + 1][1]]  # 右边框的左上角顶点
                    x_diff, y_diff = calculate_pixel_difference(right_top_point, left_top_point)
                    # # 计算每列的总和
                    # row_sums = np.sum(thresh, axis=1)
                    # print("列和:", row_sums)
                    # 总像素减白像素除以宽/高 ，得到一行/一列黑像素有多少
                    # 平均值
                    if summ:
                        midd = ((w * h) - summ) // h
                        print("狂赌像素", midd)
                        # 计算每排裂缝像素
                        # vertical_list = [c_depth_y2 - c_depth_y1 - i for i in row_sums]
                        vertical_list1 = int(midd * self.pixel_size_width_m)
                    else:
                        print(1)
                        vertical_list1 = 0
                    if self.height != 1500 or self.width != 2000:
                        vertical_list1 = vertical_list1 / 0.875
                    x_diff_mm = int(x_diff * self.pixel_size_width_m)
                    y_diff_mm = int(y_diff * self.pixel_size_height_m)
                    print(
                        f"横向 - 框 {i} 和框 {i + 1} 之间的像素差值: x差 = {x_diff_mm}, y差 = {y_diff_mm}, 裂缝宽度={vertical_list1}")

                    # 获取深度值并计算差值
                    try:
                        # depth_y = (c_depth_y2 + c_depth_y1) // 2
                        # depth_value_right_top = depth_frame1[int(c_depth_x1 + 80), int(depth_y + 50)]
                        # depth_value_left_top = depth_frame1[int(c_depth_x2 + 80), int(depth_y + 50)]
                        # if self.height != 1500 or self.width != 2000:
                        #     try:
                        #         x, y, depth_value_right_top = point_xyz_data[int(depth_y -9)][int(c_depth_x1 -13)]
                        #     except:
                        #         depth_value_right_top = 0
                        #     try:
                        #         x, y, depth_value_left_top = point_xyz_data[int(depth_y -9)][int(c_depth_x2 -13)]
                        #     except:
                        #         depth_value_left_top = 0
                        # else:
                        list1, list2 = [], []
                        for i in range(1, 10):
                            x, y, depth_value_right_top1 = point_xyz_data[int(c_depth_y1 + 50 + i)][int(c_depth_x1 + 80)]
                            if depth_value_right_top1 > 50:
                                list1.append(depth_value_right_top1)
                            x, y, depth_value_right_top2 = point_xyz_data[int(c_depth_y1 + 50 - i)][int(c_depth_x1 + 80)]
                            if depth_value_right_top2 > 50:
                                list1.append(depth_value_right_top2)
                            x, y, depth_value_left_top1 = point_xyz_data[int(c_depth_y2 + 50 + i)][int(c_depth_x2 + 80)]
                            if depth_value_left_top1 > 50:
                                list2.append(depth_value_left_top1)
                            x, y, depth_value_left_top2 = point_xyz_data[int(c_depth_y2 + 50 - i)][int(c_depth_x2 + 80)]
                            if depth_value_left_top2 > 50:
                                list2.append(depth_value_left_top2)
                        # print("x加80，y+50, 去软件里看", c_depth_x1 + 80, depth_y + 50, c_depth_x2 + 80, depth_y + 50)
                        depth_value_right_top = sum(list1) / len(list1)
                        depth_value_left_top = sum(list2) / len(list2)
                        vertical_distance = self.calculate_vertical_distance(depth_value_right_top,
                                                                             depth_value_left_top)
                        # depth_frame2 = self.device.capture_depth().data()
                        # cv2.rectangle(depth_frame2, (int(c_depth_x1 + 127), int(depth_y + 157)),
                        #               (int(c_depth_x1 + 133), int(depth_y + 163)),
                        #               (0, 0, 255), 2)
                        # cv2.rectangle(depth_frame2, (int(c_depth_x2 + 127), int(depth_y + 157)),
                        #               (int(c_depth_x2 + 133), int(depth_y + 163)),
                        #               (0, 0, 255), 2)
                        # color_image = cv2.cvtColor(depth_frame, cv2.COLOR_GRAY2BGR)

                        # cv2.imshow("depth_frame", depth_frame)
                        # cv2.imshow("depth_frame1", depth_frame1)
                    except Exception as e:
                        print(f"超出边界，错误信息{e}")
                        vertical_distance = 0

                    # print(f"depthshape{depth_frame.shape}k:{depth_value_right_top},{depth_value_left_top}")
                    print(f"横向 - 框 {i} 和框 {i + 1} 之间的深度差值: {vertical_distance}mm")
                    if y_diff_mm > 10:
                        pygame.mixer.music.load("D:/PythonCode/sql/sp.mp3")  # 替换为你的音频文件路径
                        pygame.mixer.music.play()

                    self.draw_boxes(color_frame, [inside_cuofeng_boxes[i], inside_cuofeng_boxes[i + 1]],
                                    [vertical_list1], [y_diff_mm], [vertical_distance])
                else:
                    # 纵向框
                    upper_box_right_bottom = (inside_cuofeng_boxes[i][2], inside_cuofeng_boxes[i][3])  # 上边框的右下角坐标（x,y）
                    lower_box_right_top = (
                        inside_cuofeng_boxes[i + 1][2], inside_cuofeng_boxes[i + 1][1])  # 下边框的右上角坐标（x,y）

                    # 计算x方向和y方向的差值

                    x_diff = abs(upper_box_right_bottom[0] - lower_box_right_top[0])  # 计算两个框在x方向上的差异

                    y_diff = abs(upper_box_right_bottom[1] - lower_box_right_top[1])  # 计算两个框在y方向上的差异
                    # row_sums = np.sum(thresh, axis=0)
                    # print("行和:", row_sums)
                    # 总像素减白像素除以宽/高 ，得到一行/一列黑像素有多少
                    # 平均值
                    if summ:
                        midd = ((w * h) - summ) // w
                        # 计算每排裂缝像素
                        # vertical_list = [c_depth_y2 - c_depth_y1 - i for i in row_sums]
                        vertical_list2 = int(midd * self.pixel_size_width_m)
                    else:
                        vertical_list2 = 0
                    if self.height != 1500 or self.width != 2000:
                        vertical_list2 = vertical_list2 / 0.850
                    x_diff_mm = int(x_diff * self.pixel_size_width_m)

                    y_diff_mm = int(y_diff * self.pixel_size_height_m)

                    # 转换坐标以适应深度图的分辨率
                    depth_upper_box_right_bottom = (int(upper_box_right_bottom[0] * depth_width / image_width_px),
                                                    int(upper_box_right_bottom[1] * depth_height / image_height_px))
                    depth_lower_box_right_top = (int(lower_box_right_top[0] * depth_width / image_width_px),
                                                 int(lower_box_right_top[1] * depth_height / image_height_px))
                    # 获取深度值并计算差值
                    # 获取深度值并计算差值
                    try:
                        # depth_x = (c_depth_x2 + c_depth_x1) // 2
                        # depth_value_right_top = depth_frame1[int(depth_x + 80), int(c_depth_y1 + 50)]
                        # depth_value_left_top = depth_frame1[int(depth_x + 80), int(c_depth_y2 + 50)]
                        list1, list2 = [], []
                        for i in range(1, 10):
                            x, y, depth_value_right_top1 = point_xyz_data[int(c_depth_y1 + 50)][int(c_depth_x1 + 80 + i)]
                            if depth_value_right_top1 > 50:
                                list1.append(depth_value_right_top1)
                            x, y, depth_value_right_top2 = point_xyz_data[int(c_depth_y1 + 50)][int(c_depth_x1 + 80 - i)]
                            if depth_value_right_top2 > 50:
                                list1.append(depth_value_right_top2)
                            x, y, depth_value_left_top1 = point_xyz_data[int(c_depth_y2 + 50)][int(c_depth_x2 + 80 + i)]
                            if depth_value_left_top1 > 50:
                                list2.append(depth_value_left_top1)
                            x, y, depth_value_left_top2 = point_xyz_data[int(c_depth_y2 + 50)][int(c_depth_x2 + 80 - i)]
                            if depth_value_left_top2 > 50:
                                list2.append(depth_value_left_top2)
                        depth_value_right_top = int(sum(list1) / len(list1))
                        depth_value_left_top = int(sum(list2) / len(list2))

                        # print("x加80，y+50, 去软件里看", depth_x + 80, c_depth_y1 + 50, depth_x + 80, c_depth_y2 + 50)
                        # depth_difference = abs(depth_value_right_top - depth_value_left_top)
                        depth_difference = self.calculate_vertical_distance(depth_value_right_top,
                                                                            depth_value_left_top)
                        # cv2.rectangle(depth_frame, (int(c_depth_x1 - 3), int(depth_y - 3)),
                        #               (int(c_depth_x1 + 3), int(depth_y + 3)),
                        #               (0, 0, 255), 2)
                        # cv2.rectangle(depth_frame, (int(depth_x2 - 3), int(depth_y - 3)),
                        #               (int(depth_x2 + 3), int(depth_y + 3)),
                        #               (0, 0, 255), 2)
                        print(f"depthshape{depth_frame.shape}k:{depth_value_right_top},{depth_value_left_top}")
                        # print(f"横向 - 框 {i} 和框 {i + 1} 之间的深度差值: {vertical_distance}mm")
                        print(f"纵向 - 框 {i} 和框 {i + 1} 之间的深度差值: {depth_difference}mm")
                    except Exception as e:
                        print(f"超出边界，错误信息{e}")
                        depth_difference = 0

                    print(
                        f"纵向 - 框 {i} 和框 {i + 1} 之间的像素差值: x差 = {x_diff_mm}mm, y差 = {y_diff_mm}mm,, 裂缝宽度={vertical_list2}")

                    if inside_cuofeng_boxes >= 10 or depth_difference >= 3.0:
                        self.alarmStatusChanged.emit(True)
                        print("触发报警！垂直距离大于等于三毫米或水平距离大于等于十公分")
                        # if inside_cuofeng_boxes >= 10:
                        #     print("触发报警！水平距离大于等于十公分。")
                        # elif depth_difference >= 3.0:
                        #     print("触发报警！垂直距离大于等于三毫米。")

                        pygame.mixer.music.load("D:/PythonCode/sql/cz.mp3")  # 替换为你的音频文件路径
                        pygame.mixer.music.play()
                        self.detected_label = {
                            "LC": depth_difference,
                            "box1": inside_cuofeng_boxes[i],
                            "box2": inside_cuofeng_boxes[i + 1]
                        }
                        self.detected_labels.append(self.detected_label)
                        if self.detected_labels:
                            max_lc = max(label['LC'] for label in self.detected_labels if 'LC' in label)
                            self.lcMaxUpdated.emit(f"最大LC: {max_lc}")  # 发射信号，包含最大LC值
                        else:
                            self.lcMaxUpdated.emit("未检测到任何对象")  # 如果没有标签，发射此消息
                        # 打印或处理大于3毫米的深度差的标签
                    else:
                        self.alarmStatusChanged.emit(False)
                    for label in self.detected_labels:
                        print(f"深度差大于3毫米的标签: {label}")
                    # distance_text = f"lc:{depth_difference},lf:{vertical_list2},cf:{x_diff_mm}"
                    # cv2.putText(color_frame, distance_text, (center1[0],center1[1]-10),
                    #             cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 0, 255), 2)

                    self.draw_boxes(color_frame, [inside_cuofeng_boxes[i], inside_cuofeng_boxes[i + 1]],
                                    [vertical_list2], [y_diff_mm], [depth_difference])

        self.draw_boxes_for_type(boxes_type_1, color_frame, color=(0, 255, 0))
        self.draw_boxes_for_type(cuofengg_boxes, color_frame, color=(255, 0, 0))

    def trigger_alarm(self):
        # 触发报警的操作，可以是发出警告声音、显示警告信息等
        print("触发报警！水平距离大于等于十公分。")
        pygame.mixer.music.load("D:/PythonCode/sql/sp.mp3")  # 替换为你的音频文件路径
        pygame.mixer.music.play()

    def trigger_alarm1(self):
        # 触发报警的操作，可以是发出警告声音、显示警告信息等
        print("触发报警！垂直距离大于等于三毫米。")
        pygame.mixer.music.load("D:/PythonCode/sql/cz.mp3")  # 替换为你的音频文件路径
        pygame.mixer.music.play()

    def analyse_data(self):
        # 在这里执行分析
        # 发射一个信号，通知分析完成
        self.analysisCompleted.emit()

    def get_depth(self):
        return self.depth_image

    def get_current_images(self):
        return self.current_color_image, self.current_depth_image

    def get_label_stats(self):
        return self.label_stats

    # 假设这是触发ZoomWindow显示的方法
    def on_images_captured(self, color_img, depth_img):
        # 创建ZoomWindow实例
        zoom_window = ZoomWindow(color_img, self, self.parent_window)
        # 设置detected_labels（如果已有）
        zoom_window.set_detected_labels(self.detected_labels)
        print(f'测试:', self.detected_labels)
        # 新增：传递label_stats
        zoom_window.set_label_stats(self.get_label_stats())  # 确保你有这样的方法获取label_stats
        # 新增传递模型名称
        zoom_window.set_model_name(self.selected_model)
        print('ce:', self.get_label_stats())
        zoom_window.exec()


class SerialReaderThread(QThread):
    def __init__(self, serial_port, data_handler):
        super().__init__()
        self.serial_port = serial_port
        self.data_handler = data_handler

    def run(self):
        while True:
            try:
                data = self.serial_port.read(2)  # 读取期望的数据长度
                if data:
                    self.data_handler(data)
            except Exception as e:
                # 处理读取串口数据时的异常
                print(f"读取串口数据时出错：{str(e)}")


class ZoomableImageLabel(QLabel):
    def __init__(self):
        super().__init__()
        self.setAlignment(Qt.AlignCenter)
        self.setScaledContents(False)
        self.is_zoomed = False
        self.zoom_factor = 1.0
        self.setMinimumSize(50, 50)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.setMouseTracking(True)
        self.last_mouse_pos = None

    def mouseDoubleClickEvent(self, event):
        if self.is_zoomed:
            self.setGeometry(self.original_geometry)
            self.setScaledContents(False)
            self.is_zoomed = False
            self.zoom_factor = 1.0
        else:
            self.original_geometry = self.geometry()
            self.setGeometry(self.parentWidget().geometry())
            self.setScaledContents(True)
            self.is_zoomed = True
            self.zoom_factor = 1.0
        self.update_pixmap()

    def wheelEvent(self, event):
        if event.angleDelta().y() > 0:
            self.zoom_factor *= 1.1
        else:
            self.zoom_factor /= 1.1
        self.update_pixmap()

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.last_mouse_pos = event.pos()

    def mouseMoveEvent(self, event):
        if self.last_mouse_pos:
            delta = event.pos() - self.last_mouse_pos
            self.move(self.pos() + delta)
            self.last_mouse_pos = event.pos()

    def mouseReleaseEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.last_mouse_pos = None

    def update_pixmap(self):
        if self.pixmap():
            size = self.size() * self.zoom_factor
            self.setPixmap(self.pixmap().scaled(size, Qt.KeepAspectRatio, Qt.SmoothTransformation))


class ZoomableGraphicsView(QGraphicsView):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setDragMode(QGraphicsView.ScrollHandDrag)
        self.setTransformationAnchor(QGraphicsView.AnchorUnderMouse)
        self.setResizeAnchor(QGraphicsView.AnchorUnderMouse)

    def wheelEvent(self, event):
        factor = 1.1 if event.angleDelta().y() > 0 else 0.9
        self.scale(factor, factor)


class ZoomWindow(QDialog):
    def __init__(self, image, capture_thread, parent_window):
        super().__init__()
        self.setWindowTitle("放大图像")
        self.setGeometry(100, 100, 1400, 800)
        self.image = image
        self.capture_thread = capture_thread  # 捕获线程的引用
        self.parent_window = parent_window  # MyWindow 窗口的引用
        self.detected_labels = []
        self.label_stats = {}
        self.init_ui()
        self.capture_thread.alarmStatusChanged.connect(self.update_alarm_status)

    def init_ui(self):
        self.button_style = """
                          QPushButton {
                              background-color: #89CFF0;
                              border-style: solid;
                              border-radius: 15px;
                              border-width: 2px;
                              border-color: #0057B7;
                              min-width: 90px;
                              max-width: 90px;
                              min-height: 90px;
                              max-height: 90px;
                          }
                          QPushButton:pressed {
                              background-color: #0057B7;
                              border-color: #002855;
                          }
                          QPushButton:hover {
                              border-width: 3px;
                          }
                      """
        self.palette = QPalette()  # 创建调色板
        self.setPalette(self.palette)  # 将调色板应用于窗口

        self.graphics_view = ZoomableGraphicsView(self)
        self.scene = QGraphicsScene(self)
        self.pixmap_item = QGraphicsPixmapItem(QPixmap.fromImage(self.image))
        self.scene.addItem(self.pixmap_item)
        self.graphics_view.setScene(self.scene)

        # 添加保存和取消按钮
        confirm_btn = QPushButton("停止")
        confirm_btn.setStyleSheet(self.button_style)
        gather_btn = QPushButton("继续采集")
        gather_btn.setStyleSheet(self.button_style)
        save_button = QPushButton("保存")
        save_button.setStyleSheet(self.button_style)
        cancel_button = QPushButton("取消")
        cancel_button.setStyleSheet(self.button_style)

        self.location = QLineEdit("检测位置")
        self.people = QLineEdit("巡检人员")
        self.damageDesc = QLineEdit("病害位置描述")
        self.location_lxt = QLineEdit("病害位置")
        self.description_des = QLineEdit("病害描述")
        self.disposal_dis = QLineEdit("是否处置")
        self.disposal_1 = QLineEdit("处理措施")
        # 保存按钮点击事件
        save_button.clicked.connect(self.save_image)
        # 绑定按钮点击事件
        confirm_btn.clicked.connect(self.stop_capture)
        gather_btn.clicked.connect(self.resume_capture)
        # 取消按钮点击事件
        cancel_button.clicked.connect(self.reject)

        # 创建按钮布局
        button_layout = QHBoxLayout()
        button_layout.addWidget(confirm_btn)
        button_layout.addWidget(gather_btn)
        button_layout.addWidget(save_button)
        button_layout.addWidget(cancel_button)
        label_layout = QHBoxLayout()
        label_layout.addWidget(self.location)

        label_layout.addWidget(self.people)
        label_layout.addWidget(self.location_lxt)
        label_layout.addWidget(self.description_des)
        label_layout.addWidget(self.disposal_dis)
        label_layout.addWidget(self.disposal_1)

        # 创建主布局
        layout = QVBoxLayout()
        layout.addWidget(self.graphics_view)
        layout.addLayout(button_layout)
        layout.addLayout(label_layout)

        self.setLayout(layout)

    def update_alarm_status(self, is_alarm):
        if is_alarm:
            self.palette.setColor(QPalette.Window, QColor(Qt.red))
        else:
            self.palette.setColor(QPalette.Window, QColor(Qt.white))
        self.setPalette(self.palette)
        self.repaint()  # 立即请求重绘

    def stop_capture(self):
        self.capture_thread.stopCapture.emit()  # 发射停止采集信号

    def resume_capture(self):
        self.capture_thread.resumeCapture.emit()  # 发射恢复采集信号

    def set_detected_labels(self, labels):
        self.detected_labels = labels

    def set_label_stats(self, label_stats):
        self.label_stats = label_stats

    def set_model_name(self, model_name):
        self.model_name = model_name

    def save_image(self):
        current_date = QDateTime.currentDateTime().toString("yyyyMMdd")
        folder_path = f"{current_date}"  # 只使用当前日期作为文件夹路径，不包含"./"
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        current_time = QDateTime.currentDateTime().toString("yyyy-MMdd-hhmmss")
        time_for_create_time = QDateTime.currentDateTime().toString("yyyy HH:mm:ss")
        image_file_name = f"{current_time}.png"
        text_file_name = f"{current_date}.json"

        # 获取选项中的线路和车站名称
        selected_options = self.parent_window.get_selected_options()
        line_name = selected_options["line"]
        station_name = selected_options["station"]

        # 构建Excel文件名
        excel_file_name = f"{line_name}_{station_name}.xlsx"
        excel_save_path = os.path.join(os.getcwd(), excel_file_name)  # 保存到项目根目录

        text_save_path = os.path.join(folder_path, text_file_name)
        image_save_path = os.path.join(folder_path, image_file_name)

        self.image.save(image_save_path)

        people_text = self.people.text()
        location_text1 = self.location.text()
        location_text = self.location_lxt.text()
        description_text = self.description_des.text()
        disposal_text = self.disposal_dis.text()
        disposal1_text = self.disposal_1.text()
        create_time = f"{current_date}"

        # 检查文件是否存在并且不为空
        if os.path.exists(text_save_path) and os.path.getsize(text_save_path) > 0:
            try:
                with open(text_save_path, 'r', encoding='utf-8') as file:
                    existing_data = json.load(file)
            except json.JSONDecodeError:
                existing_data = {"data": []}
        else:
            existing_data = {"data": []}

        # 新的数据条目
        new_data_entry = {
            "inspector": people_text,
            "location": location_text1,
            "line_name": line_name,
            "station_name": station_name,
            "ceiling_name": selected_options["area"],
            "depart_name": selected_options["center"],
            "create_time": create_time,
            "damageDesc": "墙面有裂缝",
            "damageType": self.label_stats,
            "damageDetail": description_text,
            "damageLocation": location_text,
            "picName": image_file_name,
            "ceilingType": self.model_name,
            "isDisposed": disposal_text,
            "measures": disposal1_text,
        }

        # 将新数据条目追加到现有数据中
        existing_data["data"].append(new_data_entry)

        # 将更新后的数据保存到文件中
        with open(text_save_path, 'w', encoding='utf-8') as file:
            json.dump(existing_data, file, ensure_ascii=False, indent=4)

        # 检查Excel文件是否存在，存在则读取，不存在则创建新的
        if os.path.exists(excel_save_path):
            workbook = load_workbook(excel_save_path)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Create Time"])

        # 检查表格中是否已经存在该文件夹名称
        create_time_exists = False
        for row in sheet.iter_rows(values_only=True):
            if row[0] == create_time:
                create_time_exists = True
                break

        # 如果不存在该文件夹名称，则将保存时间写入Excel文件
        if not create_time_exists:
            sheet.append([create_time])
            workbook.save(excel_save_path)
        self.accept()


class MyWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        # 初始化选择项
        self.selected_line = ""
        self.selected_station = ""
        self.selected_area = ""
        self.selected_center = ""
        # 获取 Dracula 颜色
        dracula = DraculaColors()
        self.initial_measurements_to_discard = 5  # 需要丢弃的初始测量次数
        self.current_measurement_count = 0  # 当前丢弃的测量次数
        self.device = Device()
        self.setWindowTitle("SDK")
        self.setGeometry(100, 100, 800, 600)
        self.data = CaptureAllData()
        self.save_images_flag = False
        self.t_name = ""
        self.num_collect = 0
        self.is_measurement_started = False  # 初始化is_measurement_started
        self.is_measurement_paused = False  # 初始化is_measurement_paused
        self.model_combo = CustomComboBox()
        self.setStyleSheet(f"""
                   QWidget {{
                       background-color: {dracula.background};
                       color: {dracula.foreground};
                   }}
                   QPushButton {{
                       background-color: {dracula.purple};
                       color: {dracula.foreground};
                       border: none;
                       padding: 20px;
                       border-radius: 20px;

                   }}
                   QPushButton:hover {{
                       background-color: {dracula.pink};
                   }}
                   QLabel {{
                       color: {dracula.cyan};
                       font-size: 15px;
                   }}
               """)
        layout = QVBoxLayout()
        btn_layout = QHBoxLayout()
        label_layout = QHBoxLayout()
        tag_layout = QHBoxLayout()
        sta_layout = QHBoxLayout()
        unit_layout = QHBoxLayout()
        work_layout = QHBoxLayout()
        image_layout = QHBoxLayout()
        alarm_layout = QHBoxLayout()
        layout.addLayout(tag_layout, 0)
        layout.addLayout(unit_layout)
        layout.addLayout(work_layout)
        layout.addLayout(alarm_layout)
        layout.addLayout(btn_layout)
        layout.addLayout(sta_layout)
        layout.addLayout(label_layout)
        # 初始化字体设置
        font = QFont("Arial", 20)  # 设置字体为Arial，大小为20

        self.depth_differences = []
        self.line_label = QLabel("选择线路")
        self.line_label.setAlignment(Qt.AlignCenter)
        self.line_label.setFrameStyle(QLabel.Panel | QLabel.Sunken)
        self.line_label.setFixedSize(130, 30)
        self.line_label.mousePressEvent = self.update_lines

        self.line_combo = QComboBox()
        self.line_combo.setFont(font)  # 设置下拉框的字体

        self.station_label = QLabel("选择车站")
        self.station_label.setAlignment(Qt.AlignCenter)
        self.station_label.setFrameStyle(QLabel.Panel | QLabel.Sunken)
        self.station_label.setFixedSize(130, 30)
        self.station_label.mousePressEvent = self.update_stations
        self.station_combo = QComboBox()
        self.station_combo.setFont(font)  # 设置下拉框的字体
        self.area_label = QLabel("选择区域")
        self.area_label.setAlignment(Qt.AlignCenter)
        self.area_label.setFrameStyle(QLabel.Panel | QLabel.Sunken)
        self.area_label.setFixedSize(130, 30)
        self.area_label.mousePressEvent = self.update_areas
        self.area_combo = QComboBox()
        self.area_combo.setFont(font)  # 设置下拉框的字体
        # 添加选项：站台、站厅、通道

        self.line_combo.setFixedSize(160, 40)  # 设置固定大小为宽度 160，高度 40
        self.station_combo.setFixedSize(160, 40)
        self.area_combo.setFixedSize(160, 40)
        tag_layout.addWidget(self.line_label, 0, Qt.AlignTop)
        tag_layout.addWidget(self.line_combo, 0, Qt.AlignTop)
        tag_layout.addWidget(self.station_label, 0, Qt.AlignTop)
        tag_layout.addWidget(self.station_combo, 0, Qt.AlignTop)
        tag_layout.addWidget(self.area_label, 0, Qt.AlignTop)
        tag_layout.addWidget(self.area_combo, 0, Qt.AlignTop)

        self.centre_label = QLabel("中心")
        self.centre_label.setAlignment(Qt.AlignCenter)
        self.centre_label.setFrameStyle(QLabel.Panel | QLabel.Sunken)
        self.centre_label.setFixedSize(130, 30)
        self.centre_label.mousePressEvent = self.update_centers

        self.center_combo = QComboBox()
        self.center_combo.setFont(font)  # 设置下拉框的字体

        unit_layout.addWidget(self.centre_label)
        unit_layout.addWidget(self.center_combo)

        self.centre_label1 = QLabel("工区")
        self.centre_label1.setAlignment(Qt.AlignCenter)
        self.centre_label1.setFrameStyle(QLabel.Panel | QLabel.Sunken)
        self.centre_label1.setFixedSize(130, 30)

        self.area_combo2 = QComboBox()
        # 添加选项：站台、站厅、通道
        area_options = ["工区1", "工区2", "工区3"]
        self.area_combo2.setFont(font)  # 设置下拉框的字体
        self.area_combo2.addItems(area_options)

        unit_layout.addWidget(self.centre_label1)
        unit_layout.addWidget(self.area_combo2)

        self.centre_label2 = QLabel("吊顶样式")
        self.centre_label2.setAlignment(Qt.AlignCenter)
        self.centre_label2.setFrameStyle(QLabel.Panel | QLabel.Sunken)
        self.centre_label2.setFixedSize(130, 30)
        self.centre_label2.mousePressEvent = self.update_ceiling_types

        self.ceiling_type_combo = QComboBox()
        self.ceiling_type_combo.setFont(font)  # 设置下拉框的字体
        self.area_combo2.setFixedSize(160, 40)  # 设置固定大小为宽度 160，高度 40
        self.center_combo.setFixedSize(160, 40)
        self.ceiling_type_combo.setFixedSize(160, 40)
        unit_layout.addWidget(self.centre_label2)
        unit_layout.addWidget(self.ceiling_type_combo)

        self.model_label = QLabel("吊顶类型")
        self.model_label.setAlignment(Qt.AlignCenter)
        self.model_label.setFrameStyle(QLabel.Panel | QLabel.Sunken)
        self.model_label.setFixedSize(130, 30)

        self.model_combo = QComboBox()
        # 向下拉框添加模型选项
        model_options = ["铝板", "栅格", "冲孔板", "勾搭铝板", "拉网", "铝条板", "栅格加网", "铝方通", "矿棉板", "垂片"]
        self.model_combo.setFixedSize(160, 40)
        self.model_combo.setFont(font)  # 设置下拉框的字体
        self.model_combo.addItems(model_options)
        self.model_combo.currentIndexChanged.connect(self.update_selected_model)
        work_layout.addWidget(self.model_label)
        work_layout.addWidget(self.model_combo)
        # 创建一个水平布局来包含选项标签和图片标签
        item_layout = QHBoxLayout()
        image_layout.addLayout(item_layout)

        # 将图片布局添加到工作布局中
        work_layout.addLayout(image_layout)
        self.capture_interval_label = QLabel("采集间隔（毫秒）")
        self.capture_interval_label.setAlignment(Qt.AlignCenter)
        self.capture_interval_label.setFrameStyle(QLabel.Panel | QLabel.Sunken)
        self.capture_interval_label.setFixedSize(130, 30)

        self.capture_interval_combo = QComboBox()
        interval_options = ["10000", "15000", "20000", "25000", "35000"]  # 采集时间间隔选项，以毫秒为单位
        self.capture_interval_combo.setFont(font)  # 设置下拉框的字体
        self.capture_interval_combo.setFixedSize(160, 40)
        self.capture_interval_combo.addItems(interval_options)
        self.capture_interval_combo.setCurrentIndex(0)  # 默认选择第一个选项，即10秒（10000毫秒）
        self.capture_interval_combo.currentIndexChanged.connect(self.update_capture_interval)

        work_layout.addWidget(self.capture_interval_label)
        work_layout.addWidget(self.capture_interval_combo)

        self.capture_interval_label1 = QLabel("曝光度")
        self.capture_interval_label1.setAlignment(Qt.AlignCenter)
        self.capture_interval_label1.setFrameStyle(QLabel.Panel | QLabel.Sunken)
        self.capture_interval_label1.setFixedSize(130, 30)
        self.exposure_combo = QComboBox()
        self.exposure_combo.setFont(font)  # 设置下拉框的字体
        exposure_options = ["40", "60", "80", "100", "120", "140", "160"]  # 曝光度选项
        self.exposure_combo.setFixedSize(160, 40)
        self.exposure_combo.addItems(exposure_options)
        self.exposure_combo.currentIndexChanged.connect(self.capture_images)
        work_layout.addWidget(self.capture_interval_label1)
        work_layout.addWidget(self.exposure_combo)

        self.alarm_label1 = QLabel("报警类型及信息")
        self.alarm_label1.setAlignment(Qt.AlignCenter)
        self.alarm_label1.setFrameStyle(QLabel.Panel | QLabel.Sunken)
        self.alarm_label1.setFixedSize(130, 30)
        alarm_layout.addWidget(self.alarm_label1)
        self.data_display2 = QLineEdit()
        self.data_display2.setFixedSize(160, 40)
        alarm_layout.addWidget(self.data_display2)

        self.db_connection = sqlite3.connect("D:\PythonCode\sql\sd.db")
        self.db_cursor = self.db_connection.cursor()
        self.image_label = ZoomableImageLabel()
        self.image_label.setFixedSize(400, 350)
        self.image_label.mouseDoubleClickEvent = lambda event: self.show_zoom_window(self.image_label)
        label_layout.addWidget(self.image_label)

        self.depth_label = ZoomableImageLabel()
        self.depth_label.setFixedSize(400, 350)
        self.depth_label.mouseDoubleClickEvent = lambda event: self.show_zoom_window(self.depth_label)
        label_layout.addWidget(self.depth_label)

        self.image_label2 = ZoomableImageLabel()
        self.image_label2.setFixedSize(400, 350)
        self.image_label2.mouseDoubleClickEvent = lambda event: self.show_zoom_window(self.image_label2)
        label_layout.addWidget(self.image_label2)

        self.data_label = QLabel("采集数量: 0")
        sta_layout.addWidget(self.data_label)

        self.line = QPushButton("启动")
        # self.line.setFixedHeight(40)  # 设置按钮的高度为 40 像素
        # self.line.setStyleSheet(self.button_style)
        self.gather = QPushButton("采集")
        # self.gather.setFixedHeight(40)
        # self.gather.setStyleSheet(self.button_style)
        self.stop_1 = QPushButton("停止")
        # self.stop_1.setFixedHeight(40)
        # self.stop_1.setStyleSheet(self.button_style)
        # self.storage = QPushButton("存储")
        # self.storage.setFixedHeight(40)
        # self.storage.setStyleSheet(self.button_style)
        # self.analyse = QPushButton("分析")
        # self.analyse.setFixedHeight(40)
        # self.analyse.setStyleSheet(self.button_style)
        self.pack = QPushButton("打包")
        self.update_btn = QPushButton("更新")
        # self.pack.setFixedHeight(40)
        # self.pack.setStyleSheet(self.button_style)

        btn_layout.addWidget(self.line)
        btn_layout.addWidget(self.gather)
        btn_layout.addWidget(self.stop_1)
        # btn_layout.addWidget(self.storage)
        # btn_layout.addWidget(self.analyse)
        btn_layout.addWidget(self.pack)
        btn_layout.addWidget(self.update_btn)

        self.line.clicked.connect(self.line_1)
        self.gather.clicked.connect(self.start_timer)
        self.stop_1.clicked.connect(self.stop_timer)
        self.update_btn.clicked.connect(self.update_data)
        # self.update_btn.connect()
        # self.storage.clicked.connect(self.save_images)
        # self.analyse.clicked.connect(self.analysis_completed)
        self.pack.clicked.connect(self.pack_1)

        layout.addLayout(btn_layout)
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_measurement_data)
        self.image_label.installEventFilter(self)
        self.depth_label.installEventFilter(self)

        self.result_images = []  # 存储分析结果图像
        self.current_result_index = 0
        self.model_combo.setCurrentIndex(0)  # 设置吊顶类型下拉框为铝板
        self.update_selected_model(0)  # 更新选中的模型，这会触发显示对应图片的功能
        self.connected = False
        self.current_station_dir = ""
        self.your_class_instance = YourClass()
        self.already_warned = False
        self.capture_thread = ImageCaptureThread(self.device, self)
        main_widget = QWidget()
        main_widget.setLayout(layout)
        self.setCentralWidget(main_widget)

        # self.received_data = []
        self.alarm_label4 = QLabel("测距仪数据")
        self.alarm_label4.setAlignment(Qt.AlignCenter)
        self.alarm_label4.setFrameStyle(QLabel.Panel | QLabel.Sunken)
        self.alarm_label4.setFixedSize(130, 30)
        alarm_layout.addWidget(self.alarm_label4)
        self.data_display = QTextEdit()
        self.data_display.setPlainText("这里显示数据")
        self.data_display.setFixedSize(160, 40)
        alarm_layout.addWidget(self.data_display)
        # self.button_2 = QPushButton("HEX")
        # self.button_2.setStyleSheet(self.button_style)
        # self.button_2.clicked.connect(self.on_read_measurement)
        # btn_layout.addWidget(self.button_2)
        self.button_3 = QPushButton("开始测距")
        # self.button_3.setStyleSheet(self.button_style)
        self.button_3.clicked.connect(self.on_start_measurement)

        # self.button_4 = QPushButton("打开测距激光")
        # self.button_4.setStyleSheet(self.button_style)
        # self.button_4.clicked.connect(self.on_laser_on)
        # btn_layout.addWidget(self.button_4)
        btn_layout.addWidget(self.button_3)
        # self.button_5 = QPushButton("关闭测距激光")
        # self.button_5.setStyleSheet(self.button_style)
        # self.button_5.clicked.connect(self.on_laser_off)
        # btn_layout.addWidget(self.button_5)
        self.button_1 = QPushButton("连接测距仪")
        # self.button_1.setStyleSheet(self.button_style)
        self.button_1.clicked.connect(self.connect_or_disconnect)
        btn_layout.addWidget(self.button_1)

        spacer = QSpacerItem(20, 40, QSizePolicy.Minimum, QSizePolicy.Expanding)
        layout.addSpacerItem(spacer)

        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_measurement_data)
        self.timer.start(1000)

        self.alarm_label3 = QLabel("选择串口")
        self.alarm_label3.setAlignment(Qt.AlignCenter)
        self.alarm_label3.setFrameStyle(QLabel.Panel | QLabel.Sunken)
        self.alarm_label3.setFixedSize(130, 30)
        alarm_layout.addWidget(self.alarm_label3)

        self.serial_port_combo = QComboBox()  # 创建一个新的下拉框用于显示串口
        self.serial_port_combo.setFixedSize(160, 40)
        self.load_serial_ports()  # 加载串口
        alarm_layout.addWidget(self.serial_port_combo)  # 将串口下拉框添加到布局中
        # 绑定下拉框变化事件
        self.line_combo.currentIndexChanged.connect(self.update_selected_options)
        self.station_combo.currentIndexChanged.connect(self.update_selected_options)
        self.area_combo.currentIndexChanged.connect(self.update_selected_options)
        self.center_combo.currentIndexChanged.connect(self.update_selected_options)

        # 省略之前的代码...

    def update_selected_options(self):
        self.selected_line = self.line_combo.currentText()
        self.selected_station = self.station_combo.currentText()
        self.selected_area = self.area_combo.currentText()
        self.selected_center = self.center_combo.currentText()

    def get_selected_options(self):
        return {
            "line": self.selected_line,
            "station": self.selected_station,
            "area": self.selected_area,
            "center": self.selected_center
        }

    def load_serial_ports(self):
        ports = list(serial.tools.list_ports.comports())
        port_names = [port.device for port in ports]
        self.serial_port_combo.clear()
        self.serial_port_combo.addItems(port_names)

    def auto_connect_measurement_device(self):
        baudrate = 19200
        data_bits = serial.EIGHTBITS
        stop_bits = serial.STOPBITS_ONE
        parity = serial.PARITY_NONE
        flow_control = False

        selected_port = self.serial_port_combo.currentText()  # 获取用户选择的串口

        try:
            self.serial_port = serial.Serial(
                port=selected_port,
                baudrate=baudrate,
                bytesize=data_bits,
                stopbits=stop_bits,
                parity=parity,
                rtscts=flow_control,
                xonxoff=flow_control
            )
            if self.serial_port.is_open:
                QMessageBox.information(self, "连接成功", f"已成功连接到串口 {selected_port}")
                self.button_1.setText("断开")
                self.button_1.clicked.disconnect()
                self.button_1.clicked.connect(self.disconnect_serial)
                self.connected = True
                self.serial_port_name = selected_port
        except Exception as e:
            print(f"无法连接到串口 {selected_port}：{str(e)}")
            QMessageBox.warning(self, "连接失败", "未找到可用的串口或连接失败")

    def connect_or_disconnect(self):
        if not self.connected:
            self.auto_connect_measurement_device()
        else:
            self.disconnect_serial()

    def disconnect_serial(self):
        if self.connected:
            try:
                self.serial_port.close()
                QMessageBox.information(self, "断开连接", "已成功断开串口连接")
                self.button_1.setText("连接")
                self.button_1.clicked.disconnect()
                self.button_1.clicked.connect(self.connect_or_disconnect)
                self.connected = False
            except Exception as e:
                QMessageBox.warning(self, "断开连接失败", f"断开串口连接时出错：{str(e)}")

    def send_command_1(self, command):
        try:
            if self.connected and self.serial_port.is_open:
                self.serial_port.write(bytes.fromhex(command))
                time.sleep(0.1)
                response = self.serial_port.read_all()
                return response
            else:
                QMessageBox.warning(self, "错误", "串口未打开，请先连接串口")
                return b''
        except Exception as e:
            QMessageBox.warning(self, "发送命令失败", f"发送命令时出错：{str(e)}")
            return b''

    def laser_on(self):
        try:
            command = "AA 00 01 BE 00 01 00 01 C1"
            self.response = self.send_command_1(command)
            print("激光已开启")
        except Exception as e:
            print("错误：", e)

    def laser_off(self):
        try:
            command = "AA 00 01 BE 00 01 00 00 C0"
            self.response = self.send_command_1(command)
            print("激光已关闭")
        except Exception as e:
            print("错误：", e)

    def start_continuous_measurement(self):
        command = "AA 00 00 20 00 01 00 04 25"
        self.response = self.send_command_1(command)

    def read_measurement_result(self):
        command = "AA 80 00 22 A2"
        response = self.send_command_1(command)
        return response

    def on_laser_on(self):
        self.laser_on()

    def on_laser_off(self):
        self.laser_off()

    def on_start_measurement(self):
        self.start_continuous_measurement()
        self.is_measurement_started = True
        self.is_measurement_paused = False
        self.current_measurement_count = 0  # 重置测量计数器

    def parse_measurement_data(self, data):
        print("测量数据字节串:", data.hex())
        distance_data = data[6:10]
        distance_int = int.from_bytes(distance_data, byteorder='big', signed=True)
        scaling_factor = 0.001
        distance_meters = distance_int * scaling_factor
        print("解析出的距离 (米):", distance_meters)
        return distance_meters

    def on_read_measurement(self):
        measurement_data = self.read_measurement_result()
        distance_meters = self.parse_measurement_data(measurement_data)
        print("测量距离 (米):", distance_meters)
        self.data_display.append(f"测量距离 (米): {distance_meters:.2f} 米")

    def update_measurement_data(self):
        try:
            if self.is_measurement_started and not self.is_measurement_paused:
                measurement_data = self.read_measurement_result()
                self.distance_meters = self.parse_measurement_data(measurement_data)

                # 检查是否需要丢弃初始测量数据
                if self.current_measurement_count < self.initial_measurements_to_discard:
                    self.current_measurement_count += 1
                    print(f"丢弃初始测量数据 {self.current_measurement_count}/{self.initial_measurements_to_discard}")
                    return

                self.data_display.append(f"测量距离 (米): {self.distance_meters:.2f} 米")

                # 使用一个小范围来比较浮点数

                tolerance = 0.05
                if self.distance_meters > 1.5 + tolerance:
                    self.your_class_instance.relay1_state_changed(True)
                    time.sleep(0.1)
                    self.your_class_instance.relay2_state_changed(True)
                    time.sleep(0.1)
                    self.your_class_instance.relay3_state_changed(False)
                    time.sleep(0.1)
                    self.your_class_instance.relay4_state_changed(False)
                elif self.distance_meters < 1.5 - tolerance:
                    self.your_class_instance.relay1_state_changed(False)
                    time.sleep(0.1)
                    self.your_class_instance.relay2_state_changed(False)
                    time.sleep(0.1)
                    self.your_class_instance.relay3_state_changed(True)
                    time.sleep(0.1)
                    self.your_class_instance.relay4_state_changed(True)
                else:
                    # 当距离在 1.5米 ± 容差 范围内时
                    self.turn_off_all_relays()
                    QMessageBox.information(self, "测量停止", "测距仪已测得1.5米，测量已停止。")
                    self.is_measurement_paused = True
                time.sleep(0.1)
        except Exception as e:
            print("更新测量数据时发生错误:", str(e))

    def turn_off_all_relays(self):
        self.your_class_instance.relay1_state_changed(False)
        time.sleep(0.1)
        self.your_class_instance.relay2_state_changed(False)
        time.sleep(0.1)
        self.your_class_instance.relay3_state_changed(False)
        time.sleep(0.1)
        self.your_class_instance.relay4_state_changed(False)

    def on_stop_measurement(self):
        self.is_measurement_started = False
        self.is_measurement_paused = False
        self.timer.stop()
        print("测量已停止。")

    def update_capture_interval(self):
        self.selected_interval = int(self.capture_interval_combo.currentText())  # 获取用户选择的采集时间间隔（毫秒）
        print("采集间隔更新为:", self.selected_interval)

    def pack_1(self):
        # 获取当前日期，格式为YYYYMMDD
        today = datetime.now().strftime('%Y%m%d')

        # 根据当前日期设置要打包的文件夹路径
        folder_path = f'D:/PycharmProjects/pythonProject/{today}'

        # 根据当前日期设置压缩文件的输出路径和名称
        zip_path = f'D:/fenlei/{today}.zip'

        # 创建一个ZIP文件
        with ZipFile(zip_path, 'w') as zipf:
            # os.walk遍历文件夹
            for root, dirs, files in os.walk(folder_path):
                for file in files:
                    # 创建文件的绝对路径
                    file_path = os.path.join(root, file)
                    # 将文件添加到zip文件中，arcname是文件在zip文件中的名称，这里使用相对路径
                    zipf.write(file_path, arcname=os.path.relpath(file_path, folder_path))

        print("ZIP文件创建成功!")

    def loadWorkAreas(self):
        try:
            self.db_cursor.execute("SELECT AreaName FROM WorkAreas ORDER BY AreaName ASC")
            work_areas = [row[0] for row in self.db_cursor.fetchall()]
            self.area_combo2.clear()  # 清除旧的选项
            self.area_combo2.addItems(work_areas)  # 加载新的选项
        except Exception as e:
            print("Error loading work areas: ", e)

    def loadInspectors(self):
        try:
            self.db_cursor.execute("SELECT Name FROM Inspectors ORDER BY Name ASC")
            inspectors = [row[0] for row in self.db_cursor.fetchall()]
            self.area_combo3.clear()  # 清除旧的选项
            self.area_combo3.addItems(inspectors)  # 加载新的选项
        except Exception as e:
            print("Error loading inspectors: ", e)

    def update_detection_result(self, detection_img):
        # 在 image_label2 中显示检测结果
        self.image_label2.setPixmap(QPixmap.fromImage(detection_img))
        self.image_label2.setAlignment(Qt.AlignCenter)

    def update_depth_label(self, depth_text):
        self.depth_label2.setText(depth_text)
        # self.calculate_and_display_difference()

    def update_depth_label1(self, depth_text1):
        self.depth_label3.setText(depth_text1)
        # self.calculate_and_display_difference()

    def update_data(self):
        selected_line = self.line_combo.currentText()
        selected_station = self.station_combo.currentText()
        if not selected_line or not selected_station:
            QMessageBox.warning(self, "选择错误", "请选择线路和车站")
            return

        excel_file_name = f"{selected_line}_{selected_station}.xlsx"
        excel_file_path = os.path.join(os.getcwd(), excel_file_name)

        if not os.path.exists(excel_file_path):
            QMessageBox.warning(self, "文件不存在", f"Excel文件 {excel_file_name} 不存在")
            return

        workbook = load_workbook(excel_file_path)
        sheet = workbook.active
        latest_create_time = sheet.cell(sheet.max_row, 1).value

        self.db_cursor.execute("SELECT create_time FROM ceiling_info WHERE line_name=? AND station_name=?",
                               (selected_line, selected_station))
        db_create_time = self.db_cursor.fetchone()[0]

        if latest_create_time != db_create_time:
            json_folder_path = latest_create_time.split(" /")[0]
            json_file_path = os.path.join(json_folder_path, f"{json_folder_path}.json")

            if os.path.exists(json_file_path):
                with open(json_file_path, 'r', encoding='utf-8') as file:
                    json_data = json.load(file)

                updated = False
                if "data" in json_data:
                    for entry in json_data["data"]:
                        if (entry.get("line_name") == selected_line and
                                entry.get("station_name") == selected_station and
                                entry.get("ceiling_name") == self.selected_area):
                            entry["create_time"] = latest_create_time
                            updated = True

                            self.db_cursor.execute("""
                                UPDATE ceiling_info
                                SET line_name=?, station_name=?, ceiling_name=?, depart_name=?, create_time=?
                                WHERE line_name=? AND station_name=? AND ceiling_name=?
                            """, (
                                entry["line_name"], entry["station_name"], entry["ceiling_name"], entry["depart_name"],
                                entry["create_time"],
                                selected_line, selected_station, self.selected_area
                            ))
                            self.db_connection.commit()
                        else:
                            continue

                if updated:
                    with open(json_file_path, 'w', encoding='utf-8') as file:
                        json.dump(json_data, file, ensure_ascii=False, indent=4)

                    QMessageBox.information(self, "更新成功", "数据已更新")
                else:
                    QMessageBox.warning(self, "未更新", "没有找到匹配的数据进行更新")
            else:
                QMessageBox.warning(self, "文件不存在", f"JSON文件 {json_file_path} 不存在")
        else:
            QMessageBox.information(self, "无变化", "数据没有变化，不需要更新")

    def capture_images(self):
        # 获取用户选择的曝光时间
        exposure_time = float(self.exposure_combo.currentText())

        # 设置曝光时间
        Common.show_error(self.device.set_scan_2d_exposure_time(exposure_time))
        # 捕获彩色图像和深度图像
        self.color_image = self.device.capture_color().data()
        self.depth_image = self.device.capture_depth().data()

    def show_zoom_window(self, label):
        pixmap = label.pixmap()
        if pixmap:
            zoom_window = ZoomWindow(pixmap)
            zoom_window.exec()

    def show_zoom_window_method(self):
        zoom_window = ZoomWindow(self)
        zoom_window.exec_()

    def update_lines(self, event):
        self.db_cursor.execute("SELECT DISTINCT line_name FROM ceiling_info")
        lines = [row[0] for row in self.db_cursor.fetchall()]

        self.line_combo.clear()
        self.line_combo.addItems(lines)
        self.line_combo.show()

        self.save_dir = os.getcwd() + "/data/"
        if not os.path.exists(self.save_dir):
            os.makedirs(self.save_dir)

    def create_folder(self):
        selected_line = self.line_combo.currentText()
        selected_station = self.station_combo.currentText()
        selected_area = self.area_combo.currentText()
        if selected_line and selected_station and selected_area:
            self.current_station_dir = os.path.join(self.save_dir, selected_line, selected_station, selected_area)
            if not os.path.exists(self.current_station_dir):
                os.makedirs(self.current_station_dir)

            print(f"已创建目录：{self.current_station_dir}")
        else:
            print("请选择线路、车站和区域以创建文件夹。")

    def update_stations(self, event):
        selected_line = self.line_combo.currentText()

        if selected_line:
            self.db_cursor.execute("SELECT DISTINCT station_name FROM ceiling_info WHERE line_name=?", (selected_line,))
            stations = [row[0] for row in self.db_cursor.fetchall()]

            self.station_combo.clear()
            self.station_combo.addItems(stations)
            self.station_combo.show()

    def update_areas(self, event):
        selected_line = self.line_combo.currentText()
        selected_station = self.station_combo.currentText()
        if selected_line and selected_station:
            self.db_cursor.execute(
                "SELECT DISTINCT ceiling_name FROM ceiling_info WHERE line_name=? AND station_name=?",
                (selected_line, selected_station))
            areas = [row[0] for row in self.db_cursor.fetchall()]
            self.area_combo.clear()
            self.area_combo.addItems(areas)
            self.area_combo.show()

    def update_centers(self, event):
        selected_line = self.line_combo.currentText()
        selected_station = self.station_combo.currentText()
        selected_area = self.area_combo.currentText()
        if selected_line and selected_station and selected_area:
            self.db_cursor.execute(
                "SELECT DISTINCT depart_name FROM ceiling_info WHERE line_name=? AND station_name=? AND ceiling_name=?",
                (selected_line, selected_station, selected_area))
            centers = [row[0] for row in self.db_cursor.fetchall()]
            self.center_combo.clear()
            self.center_combo.addItems(centers)
            self.center_combo.show()

    def update_ceiling_types(self, event):
        selected_line = self.line_combo.currentText()
        selected_station = self.station_combo.currentText()
        selected_area = self.area_combo.currentText()
        if selected_line and selected_station and selected_area:
            self.db_cursor.execute(
                "SELECT DISTINCT ceiling_texture FROM ceiling_info WHERE line_name=? AND station_name=? AND ceiling_name=?",
                (selected_line, selected_station, selected_area))
            ceiling_types = [row[0] for row in self.db_cursor.fetchall()]
            self.ceiling_type_combo.clear()
            self.ceiling_type_combo.addItems(ceiling_types)
            self.ceiling_type_combo.show()

    def start_timer(self):
        self.capture_thread = ImageCaptureThread(self.device, self)

        # self.capture_thread = ImageCaptureThread(self.device)  # 重新创建线程实例
        selected_model = self.model_combo.currentText()
        self.capture_thread.set_selected_model(selected_model)
        self.capture_thread.imagesCaptured.connect(self.update_labels)
        self.capture_thread.imagesCapturedAndSaved.connect(self.save_images)
        self.capture_thread.detectionCompleted.connect(self.update_detection_result)
        self.capture_thread.depthvalue.connect(self.update_depth_label)
        self.capture_thread.depthvalue1.connect(self.update_depth_label1)
        # self.capture_thread.set_capture_interval(self.selected_interval)  # 将采集时间间隔设置到图像采集线程中
        self.capture_thread.set_capture_interval(self.selected_interval if hasattr(self,
                                                                                   'selected_interval') else 10000)  # 将采集时间间隔设置到图像采集线程中，默认值为10000毫秒（10秒）
        self.capture_thread.lcMaxUpdated.connect(self.update_alarm_label)
        self.capture_thread.start()  # 启动线程

    def update_alarm_label(self, text):
        self.data_display2.setText(text)  # 更新UI显示最大LC值

    def update_selected_model(self, model_index):
        # 获取当前选择的吊顶类型
        self.capture_thread = ImageCaptureThread(self.device, self)
        selected_model = self.model_combo.itemText(model_index)
        self.capture_thread.set_selected_model(selected_model)

    def update_depth_image(self):
        depth_min = 725  # 期望的深度范围最小值
        depth_max = 856  # 期望的深度范围最大值
        slider_value = self.depth_slider.value()
        processed_depth_image = self.capture_thread.get_depth()  # 自定义的深度图像处理函数

        if processed_depth_image is not None and isinstance(processed_depth_image, np.ndarray):
            # 将slider_value映射到725到856的范围
            depth_value = int(np.interp(slider_value, [self.depth_slider.minimum(), self.depth_slider.maximum()],
                                        [depth_min, depth_max]))

            # 映射深度图像范围到725到856
            processed_depth_image = np.clip(processed_depth_image, depth_min, depth_max)

            # 将深度图像转换为颜色图像
            depth_colormap = cv2.applyColorMap(cv2.convertScaleAbs(processed_depth_image, alpha=0.03), cv2.COLORMAP_JET)

            # 更新self.depth_label的Pixmap，显示更新后的深度图像
            depth_img = QImage(depth_colormap.data, depth_colormap.shape[1], depth_colormap.shape[0],
                               QImage.Format_RGB888)
            self.depth_label.setPixmap(QPixmap.fromImage(depth_img))
        else:
            print("Error: Invalid processed_depth_image")

    def update_labels(self, color_img, depth_img):
        self.image_label.setPixmap(QPixmap.fromImage(color_img))
        self.depth_label.setPixmap(QPixmap.fromImage(depth_img))

    def line_1(self):
        selected_line = self.line_combo.currentText()
        selected_station = self.station_combo.currentText()
        selected_area = self.area_combo.currentText()
        if selected_line and selected_station and selected_area:
            self.current_station_dir = os.path.join(self.save_dir, selected_line, selected_station, selected_area)
            if not os.path.exists(self.current_station_dir):
                os.makedirs(self.current_station_dir)

            print(f"已创建目录：{self.current_station_dir}")
        else:
            print("请选择线路、车站和区域以创建文件夹。")

        Common.find_camera_list(self)

        if Common.choose_camera_and_connect(self):
            self.connected = True
            print("Connected to the Mech-Eye device successfully.")
            QMessageBox.information(self, "连接成功", f"已成功连接到相机 ")
        else:
            print("Failed to connect to the Mech-Eye device.")
            QMessageBox.warning(self, "连接失败", f"连接到相机失败 ")

    def capture_color_map1(self):
        color_map = self.device.capture_color()
        color_file = os.path.join(self.current_station_dir, self.t_name + "ColorMap.png")
        cv2.imencode('.png', color_map.data())[1].tofile(color_file)
        print("Capture and save color image : {}".format(color_file))

    def capture_depth_map1(self):
        depth_map = self.device.capture_depth()
        depth_file = os.path.join(self.current_station_dir, self.t_name + "DepthMap.tiff")
        cv2.imencode('.tiff', depth_map.data())[1].tofile(depth_file)
        print("Capture and save depth image : {}".format(depth_file))

    def capture_point_cloud(self):
        points_xyz = self.device.capture_point_xyz()
        points_xyz_data = points_xyz.data()
        points_xyz_o3d = o3d.geometry.PointCloud()
        points_xyz_o3d.points = o3d.utility.Vector3dVector(points_xyz_data.reshape(-1, 3) * 0.001)
        o3d.io.write_point_cloud(os.path.join(self.current_station_dir, self.t_name + "PointCloudXYZ.ply"),
                                 points_xyz_o3d)
        print("Point cloud saved to path" + self.current_station_dir + self.t_name + "PointCloudXYZ.ply")

    def capture_color_point_cloud(self):
        points_xyz_bgr = self.device.capture_point_xyz_bgr().data()
        points_reshape = points_xyz_bgr.reshape(-1, 6)
        points_xyz_rgb_points = points_reshape[:, :3] * 0.001
        point_xyz_rgb_colors = points_reshape[:, 3:6][:, ::-1] / 255

        points_xyz_rgb_o3d = o3d.geometry.PointCloud()
        points_xyz_rgb_o3d.points = o3d.utility.Vector3dVector(points_xyz_rgb_points.astype(np.float64))
        points_xyz_rgb_o3d.colors = o3d.utility.Vector3dVector(point_xyz_rgb_colors.astype(np.float64))
        o3d.io.write_point_cloud(os.path.join(self.current_station_dir, self.t_name + "PointCloudXYZRGB.ply"),
                                 points_xyz_rgb_o3d)
        print("Point cloud saved to path" + self.current_station_dir + self.t_name + "PointCloudXYZ.ply")

    def save_images(self):
        if not self.connected:
            print("请先点击启动按钮连接到 Mech-Eye 设备")
            return

        self.num_collect = self.num_collect + 1
        self.t_name = time.strftime('%Y-%m-%d_%H-%M-%S', time.localtime())
        self.data_label.setText(f"采集数量: {self.num_collect}")
        print("采集一次")
        print("第%d次数据采集" % self.num_collect)

        self.create_folder()

        # 使用传递的图像保存
        self.save_color_image(self.capture_thread.color_img)
        self.save_depth_image(self.capture_thread.depth_img)

    def save_color_image(self, color_img):
        color_file = os.path.join(self.current_station_dir, self.t_name + "ColorImage.png")
        color_img.save(color_file)
        print("Capture and save color image:", color_file)

    def save_depth_image(self, depth_img):
        depth_file = os.path.join(self.current_station_dir, self.t_name + "DepthImage.png")

        depth_img.save(depth_file)
        print("Capture and save depth image:", depth_file)

    def save_ply_files(self):
        if not self.connected:
            print("请先点击启动按钮连接到 Mech-Eye 设备")
            return

        # 保存点云PLY文件
        points_xyz_ply = o3d.io.read_point_cloud(
            os.path.join(self.current_station_dir, self.t_name + "PointCloudXYZ.ply"))
        points_xyz_ply_path = os.path.join(self.current_station_dir, self.t_name + "PointCloudXYZ.ply")
        o3d.io.write_point_cloud(points_xyz_ply_path, points_xyz_ply)

        # 保存带颜色的点云PLY文件
        points_xyz_rgb_ply = o3d.io.read_point_cloud(
            os.path.join(self.current_station_dir, self.t_name + "PointCloudXYZRGB.ply"))
        points_xyz_rgb_ply_path = os.path.join(self.current_station_dir, self.t_name + "PointCloudXYZRGB.ply")
        o3d.io.write_point_cloud(points_xyz_rgb_ply_path, points_xyz_rgb_ply)

    def stop_timer(self):
        if self.capture_thread.isRunning():
            self.capture_thread.is_running = False  # 通知线程停止运行
            self.capture_thread.terminate()  # 强制终止线程
            self.capture_thread.wait()  # 等待线程完全停止
            self.timer.stop()  # 停止Qt计时器


def main():
    import sys
    from PySide6.QtWidgets import QApplication
    app = QApplication(sys.argv)
    window = MyWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
